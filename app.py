import streamlit as st
import sqlite3
import pandas as pd
import io
import openpyxl
import random
import plotly.express as px
import plotly.graph_objects as go
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, date
import logging
from dataclasses import replace

from config import DB_FILE, PRODUCTION_MODE
from core.pricing_engine import PricingEngine, PricingInput
from core.hierarchy_resolver import HierarchyResolver
from core.validators import DataSanitizer

logging.basicConfig(level=logging.WARNING)

# Impostazione pagina
st.set_page_config(page_title="Bunker Commerciale - Salov", layout="wide", initial_sidebar_state="expanded")

# ==========================================
# FUNZIONI GLOBALI E COSTANTI DI SUPPORTO
# ==========================================
OPERATIVE_COLS = [
    '[N+1] Volumi', '[N] Listino €', '[N+1] Listino €', 
    '[N+1] Sc. Fattura %', '[N+1] Contratto %'
]

def fmt_it(val, decimals=2, is_euro=False, is_pct=False, sign=False):
    if pd.isna(val): return ""
    if sign:
        s = f"{val:+,.{decimals}f}"
    else:
        s = f"{val:,.{decimals}f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    if is_euro: return f"€ {s}"
    if is_pct: return f"{s} %"
    return s

def safe_dec(val):
    if pd.isna(val) or val is None or str(val).strip() == "":
        return Decimal("0.00")
    try:
        return Decimal(str(val))
    except:
        return Decimal("0.00")

def get_subcat(row):
    """
    Classifica i prodotti in sub-categorie merceologiche partendo
    dal tipo di olio e dalla descrizione commerciale SAP.
    """
    desc = str(row['descrizione_commerciale']).upper()
    tipo = str(row['tipo_olio']).upper()
    if tipo == 'EXTRAVERGINE':
        if '100% ITA' in desc or '100%I' in desc or 'TOSC' in desc: return 'Extravergini Italiani'
        if 'BIO' in desc: return 'Extravergini Biologici'
        return 'Extravergini Comunitari'
    elif tipo == 'OLIVA': 
        return 'Olio Raffinato'
    elif tipo == 'SEMI':
        if 'ARACHIDE' in desc: return 'Semi di Arachide'
        if 'MAIS' in desc: return 'Semi di Mais'
        if 'GIRAS' in desc: return 'Semi di Girasole'
        if 'FRITT' in desc or 'FRIMX' in desc: return 'Oli per Frittura Specifici'
        if 'VINACC' in desc: return 'Semi di Vinacciolo'
        return 'Oli di Semi'
    elif tipo == 'ACETO': 
        return 'Aceto Balsamico'
    return 'Altro'

def render_badge(level):
    """
    Genera un badge HTML elegante e coordinato con la palette di colori 
    per identificare immediatamente l'ereditarietà contrattuale nel simulatore.
    """
    lvl = str(level).upper().strip()
    if lvl == "GRUPPO":
        return '<span class="badge badge-gruppo">GRUPPO (Quadro Gen.)</span>'
    elif lvl == "SOTTOGRUPPO":
        return '<span class="badge badge-sottogruppo">SOTTOGRUPPO (Consorzio)</span>'
    elif lvl == "CATEGORIA":
        return '<span class="badge badge-categoria">CATEGORIA (Famiglia)</span>'
    elif lvl == "INSEGNA":
        return '<span class="badge badge-insegna">INSEGNA LOCALE</span>'
    elif lvl == "REFERENZA":
        return '<span class="badge badge-referenza">REFERENZA (SKU)</span>'
    else:
        return f'<span class="badge badge-nessuno">{lvl}</span>'

# ==========================================
# CSS AVANZATO (CON REGOLE DI SPAZIATURA)
# ==========================================
st.markdown("""
<style>
    /* 1. IMPORTAZIONE DEI FONT AZIENDALI */
    @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;600;700&family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

    /* 2. REGOLE STRUTTURALI E COLORI DI SFONDO */
    .block-container { 
        padding-top: 2rem !important; 
        padding-bottom: 2rem !important; 
        max-width: 95% !important; 
    }
    
    /* Sfondo principale crema "Natural BG" */
    .stApp, .stApp [data-testid="stAppViewContainer"] { 
        background-color: #F7F7F2 !important; 
        color: #2D3227 !important;
        font-family: 'Inter', sans-serif !important;
    } 

    /* Sidebar con tonalità grigio-caldo */
    section[data-testid="stSidebar"] { 
        background-color: #F0F0E8 !important; 
        border-right: 1px solid #E2E2D8 !important; 
    }
    section[data-testid="stSidebar"] .stMarkdown, section[data-testid="stSidebar"] label {
        color: #2D3227 !important;
    }

    /* 3. TIPOGRAFIA */
    h1, h2, h3 { 
        color: #2D3227 !important; 
        font-family: 'Space Grotesk', sans-serif !important; 
        font-weight: 700 !important; 
        letter-spacing: -0.03em !important; 
    }
    h4, h5, h6 { 
        color: #5A6340 !important; 
        font-family: 'Space Grotesk', sans-serif !important;
        font-weight: 600 !important; 
    }
    
    /* Numeri e codici in font monospazio */
    .font-mono, td, .stMetric div[data-testid="stMetricValue"] {
        font-family: 'JetBrains Mono', monospace !important;
    }

    /* 4. METRICHE E CARTE (Componenti "Card") */
    div[data-testid="stMetric"] { 
        background-color: #FFFFFF !important; 
        padding: 20px !important; 
        border-radius: 16px !important; 
        border: 1px solid #E2E2D8 !important; 
        box-shadow: 0 1px 3px rgba(45,50,39,0.05) !important; 
    }
    div[data-testid="stMetricValue"] { 
        color: #2D3227 !important; 
        font-weight: 700 !important; 
        font-size: 1.8rem !important; 
    }
    div[data-testid="stMetricLabel"] { 
        font-size: 0.75rem !important; 
        color: #7A7E72 !important; 
        text-transform: uppercase !important; 
        letter-spacing: 0.1em !important; 
        font-weight: 700 !important; 
    }

    /* Contenitori generici (st.container con border=True) */
    div[data-testid="stVerticalBlockBorderWrapper"] { 
        background-color: #FFFFFF !important; 
        border-radius: 16px !important; 
        border: 1px solid #E2E2D8 !important; 
        box-shadow: 0 1px 4px rgba(0,0,0,0.02) !important; 
        padding: 20px !important; 
    }

    /* 5. PULSANTI (Stile primario Verde Oliva) */
    .stButton>button[kind="primary"] { 
        background-color: #5A6340 !important; 
        color: #FFFFFF !important; 
        font-family: 'Space Grotesk', sans-serif !important;
        font-weight: 700 !important; 
        border-radius: 12px !important; 
        border: none !important; 
        padding: 0.6rem 1.2rem !important; 
        box-shadow: 0 2px 4px rgba(90,99,64,0.15) !important;
        transition: all 0.2s ease-in-out !important;
    }
    .stButton>button[kind="primary"]:hover { 
        background-color: #495033 !important; 
        transform: translateY(-1px) !important;
    }
    
    /* Pulsanti secondari */
    .stButton>button[kind="secondary"] { 
        background-color: #FFFFFF !important; 
        color: #2D3227 !important; 
        border: 1px solid #E2E2D8 !important;
        border-radius: 12px !important;
        font-family: 'Space Grotesk', sans-serif !important;
        font-weight: 600 !important;
        transition: all 0.2s ease-in-out !important;
    }
    .stButton>button[kind="secondary"]:hover { 
        background-color: #F0F0E8 !important; 
        border-color: #7A7E72 !important;
    }

    /* 6. TAB DI NAVIGAZIONE (Stile minimale a pillola) */
    button[data-baseweb="tab"] {
        font-family: 'Space Grotesk', sans-serif !important;
        font-weight: 700 !important;
        font-size: 0.85rem !important;
        color: #7A7E72 !important;
        border: none !important;
        padding: 8px 16px !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: #5A6340 !important;
        background-color: #DDE2C6 !important;
        border-radius: 8px !important;
    }

    /* 7. SCROLLBARS PERSONALIZZATE */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    ::-webkit-scrollbar-track {
        background: #F7F7F2;
    }
    ::-webkit-scrollbar-thumb {
        background: #E2E2D8;
        border-radius: 4px;
    }
    ::-webkit-scrollbar-thumb:hover {
        background: #7A7E72;
    }

    /* 8. BOX DI ALLERTA MODIFICATI */
    .alert-box { 
        padding: 12px 18px; 
        border-radius: 12px; 
        font-size: 0.85rem; 
        margin-bottom: 12px; 
        border-left: 4px solid; 
    }
    .alert-warning { background-color: #FAF2F0; border-color: #A34A3F; color: #4D1A16; } 
    .alert-info { background-color: #F0F0E8; border-color: #8A9A5B; color: #2D3227; }
    .alert-success { background-color: #E9F2E9; border-color: #5A6340; color: #202418; }
    .alert-danger { background-color: #FAF2F0; border-color: #A34A3F; color: #4D1A16; }

    /* 9. BADGES DI LIVELLO RISOLTO */
    .badge {
        display: inline-block;
        padding: 4px 10px;
        border-radius: 8px;
        font-size: 0.75rem;
        font-weight: 700;
        font-family: 'Space Grotesk', sans-serif !important;
        text-transform: uppercase;
        letter-spacing: 0.03em;
        text-align: center;
    }
    .badge-gruppo { background-color: #E9F2E9 !important; color: #202418 !important; border: 1px solid #C4CCA0; }
    .badge-sottogruppo { background-color: #F0F0E8 !important; color: #5A6340 !important; border: 1px solid #E2E2D8; }
    .badge-categoria { background-color: #EAEAE0 !important; color: #2D3227 !important; border: 1px solid #D6D6CC; }
    .badge-insegna { background-color: #FAF2F0 !important; color: #8A3830 !important; border: 1px solid #EAD0CC; }
    .badge-referenza { background-color: #DDE2C6 !important; color: #495033 !important; border: 1px solid #C4CCA0; }
    .badge-nessuno { background-color: #E2E2D8 !important; color: #7A7E72 !important; border: 1px solid #D6D6CC; }

    /* 10. PREVENZIONE DEL CONFLITTO DI CONTRASTO NEI WIDGET NATIVI */
    /* Forza il testo scuro e lo sfondo chiaro nei moduli di immissione e tendine */
    div[data-baseweb="select"] > div, 
    div[data-baseweb="input"] > div,
    input, 
    select, 
    textarea {
        background-color: #FFFFFF !important;
        color: #2D3227 !important;
        -webkit-text-fill-color: #2D3227 !important; /* Forza iOS Safari */
    }
    
    /* Forza il colore scuro per i placeholder e le etichette dei campi */
    label[data-testid="stWidgetLabel"] p, label[data-testid="stWidgetLabel"] span {
        color: #2D3227 !important;
    }
    
    /* Popover dei menu a tendina (Selezionatori) */
    div[data-baseweb="popover"] *, div[role="listbox"] * {
        background-color: #FFFFFF !important;
        color: #2D3227 !important;
    }
    
    /* Allineamento dei testi interni dei componenti st.expander */
    div[data-testid="stExpander"] * {
        color: #2D3227 !important;
    }

    /* FORZA IL TESTO DELLE OPZIONI DEI RADIO BUTTONS (st.radio) E CHECKBOX (st.checkbox) A RESTARE SCURO */
    div[data-testid="stRadio"] *, 
    div[data-baseweb="radio"] *, 
    label[data-baseweb="radio"] *,
    div[data-testid="stCheckbox"] *,
    div[data-baseweb="select"] *,
    div[class*="select"] * {
        color: #2D3227 !important;
        -webkit-text-fill-color: #2D3227 !important; /* Forza iOS Safari */
    }
</style>
""", unsafe_allow_html=True)

# ==========================================
# INIZIALIZZAZIONE DATABASE E MERGE CONTRATTI
# ==========================================
def init_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute("CREATE TABLE IF NOT EXISTS anagrafica_master (ean TEXT PRIMARY KEY, codice_sap TEXT, tipo_olio TEXT, descrizione_sap TEXT, descrizione_commerciale TEXT, formato_lt REAL, confezione TEXT, pezzi_cartone INTEGER, cartoni_strato INTEGER, strati_pallet INTEGER, cartoni_pallet INTEGER, conservazione_mesi INTEGER, shelf_life_mesi INTEGER)")
    cursor.execute("CREATE TABLE IF NOT EXISTS guardrail_aziendali (ean TEXT PRIMARY KEY, min_net_net_g REAL DEFAULT 0.0)")
    cursor.execute("CREATE TABLE IF NOT EXISTS clienti (id INTEGER PRIMARY KEY AUTOINCREMENT, gruppo_macro TEXT, sottogruppo TEXT, associato_insegna TEXT, attivo BOOLEAN DEFAULT 1, UNIQUE(gruppo_macro, sottogruppo, associato_insegna))")
    cursor.execute("CREATE TABLE IF NOT EXISTS struttura_gdo (id INTEGER PRIMARY KEY AUTOINCREMENT, gruppo_macro TEXT, sottogruppo TEXT, associato_insegna TEXT, attivo BOOLEAN DEFAULT 0, UNIQUE(gruppo_macro, sottogruppo, associato_insegna))")
    cursor.execute("CREATE TABLE IF NOT EXISTS accordi_commerciali (id INTEGER PRIMARY KEY AUTOINCREMENT, gruppo_macro TEXT, sottogruppo TEXT, associato_insegna TEXT, livello TEXT, chiave_livello TEXT, listino_r REAL, sconto_1 REAL, sconto_2 REAL, sconto_3 REAL, sconto_4 REAL, sconto_5 REAL, sconto_6 REAL, sconto_7 REAL, sconto_y REAL, sconto_carico REAL, sconto_pagamento REAL, voce_contratto_1 REAL, voce_contratto_2 REAL, voce_contratto_3 REAL, voce_contratto_4 REAL, voce_contratto_5 REAL, note_locali TEXT, UNIQUE(gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello))")
    cursor.execute("CREATE TABLE IF NOT EXISTS storico_promo (id INTEGER PRIMARY KEY AUTOINCREMENT, data_salvataggio TIMESTAMP DEFAULT CURRENT_TIMESTAMP, stato_promo TEXT, gruppo_macro TEXT, sottogruppo TEXT, associato_insegna TEXT, ean TEXT, descrizione_commerciale TEXT, listino_r REAL, sconto_y REAL, sconto_z REAL, sconto_aa REAL, net_net_am REAL, volumi_stimati INTEGER, contributo_fisso REAL, contributo_pezzo REAL, costo_totale_extra REAL, note TEXT, sell_in_dal DATE, sell_in_al DATE, sell_out_dal DATE, sell_out_al DATE, min_net_net_g REAL, net_net_post_promo REAL)")
    conn.commit()

    migrazioni = [
        "ALTER TABLE storico_promo ADD COLUMN sell_in_dal DATE",
        "ALTER TABLE storico_promo ADD COLUMN sell_in_al DATE",
        "ALTER TABLE storico_promo ADD COLUMN sell_out_dal DATE",
        "ALTER TABLE storico_promo ADD COLUMN sell_out_al DATE",
        "ALTER TABLE storico_promo ADD COLUMN min_net_net_g REAL",
        "ALTER TABLE storico_promo ADD COLUMN net_net_post_promo REAL",
        "ALTER TABLE accordi_commerciali ADD COLUMN note_locali TEXT",
        "ALTER TABLE struttura_gdo ADD COLUMN sottogruppo TEXT"
    ]
    for query in migrazioni:
        try:
            cursor.execute(query)
            conn.commit()  # ✅ Se funziona, salva subito questa modifica
        except sqlite3.OperationalError:
            conn.rollback()  # 🛡️ Se fallisce (colonna già esistente), resetta lo stato ed evita il blocco!
    conn.commit()

    cursor.execute("SELECT COUNT(*) FROM anagrafica_master")
    if cursor.fetchone()[0] == 0:
        seed_baseline_data(conn)

    conn.close()

def get_listino_strutturale(conn, gruppo, sottogruppo, ean):
    """
    Esegue una ricerca gerarchica del listino_r per una data referenza 
    a livello nazionale se non è stato definito a livello locale.
    """
    cursor = conn.cursor()
    cursor.execute("""
        SELECT listino_r FROM accordi_commerciali
        WHERE UPPER(TRIM(gruppo_macro)) = UPPER(TRIM(?))
          AND (UPPER(TRIM(sottogruppo)) = UPPER(TRIM(?)) OR sottogruppo = '' OR sottogruppo IS NULL)
          AND livello = 'REFERENZA'
          AND chiave_livello = ?
          AND listino_r IS NOT NULL
        ORDER BY 
            CASE WHEN SOTTOGRUPPO IS NOT NULL AND SOTTOGRUPPO != '' THEN 1 ELSE 2 END ASC
        LIMIT 1
    """, (gruppo, sottogruppo, ean))
    row = cursor.fetchone()
    if row:
        return Decimal(str(row[0]))
    return None

def get_merged_contract(conn, gruppo, sottogruppo, insegna, ean, categoria):
    """
    Risolve il contratto fondendo gli accordi Nazionali (Strutturali) con quelli Locali (Promo).
    Garantisce che Listino, S1-S5 e PFA provengano dalla sede centrale, mentre S6, S7 e Y dall'insegna.
    """
    base = HierarchyResolver.resolve(conn, gruppo, sottogruppo, "", ean, categoria)
    
    if insegna and str(insegna).strip() != "":
        loc = HierarchyResolver.resolve(conn, gruppo, sottogruppo, insegna, ean, categoria)
        
        strutturali_attrs = [
            'listino_r', 'sconto_1', 'sconto_2', 'sconto_3', 'sconto_4', 'sconto_5', 
            'sconto_carico', 'sconto_pagamento', 
            'voce_i', 'voce_ii', 'voce_iii', 'voce_iv', 'voce_v'
        ]
        for attr in strutturali_attrs:
            if getattr(loc, attr, None) is None:
                setattr(loc, attr, getattr(base, attr, None))
                
        return loc
    return base
    
def seed_baseline_data(conn):
    cursor = conn.cursor()
    
    # 🛡️ Resetta preventivamente la transazione per ripulire eventuali errori precedenti
    conn.rollback()
    
    # Esegue le cancellazioni proteggendo ogni singola operazione
    for tabella in ["accordi_commerciali", "clienti", "anagrafica_master", "guardrail_aziendali", "struttura_gdo"]:
        try:
            cursor.execute(f"DELETE FROM {tabella}")
            conn.commit()
        except sqlite3.OperationalError:
            conn.rollback()
    
    prodotti_salov = [
        ("8002210111110", "10002713", "EXTRAVERGINE", "SAGRA EXV BOT W12x1L CLASS IT", "Ex.v. Sagra Classico lt.1", 1.0, 10.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210133440", "10003255", "EXTRAVERGINE", "SAGRA EXV 100%R-PET V12x750ML IT", "Ex.v. Sagra lt.0,75 PET", 0.75, 7.50, "Pet.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210115088", "10002716", "EXTRAVERGINE", "SAGRA GRAND EXV BOT W12x1L", "Ex.v. Sagra Grandulivo lt.1", 1.0, 10.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210127562", "10002719", "EXTRAVERGINE", "SAGRA T.VIVE EXV BOT W 12x1L", "Ex.v. Sagra Terre Vive lt.1", 1.0, 10.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210119543", "10000536", "EXTRAVERGINE", "SAGRA PROF. EXV PET C2x5L IT", "Ex.v. Sagra Prof Lt.5", 5.0, 50.00, "Pet lt 5", 2, 17, 4, 68, 14, 9),
        ("8002210112827", "10002714", "EXTRAVERGINE", "SAGRA EXV 100%I BSA BOT W12x1L IT", "Ex.v. Sagra Bassa Acidità 100% ITA lt.1", 1.0, 15.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210127425", "10002715", "EXTRAVERGINE", "SAGRA EXV 100%I BOT W 12x1L", "Ex.v. Sagra 100% Italiano lt.1", 1.0, 15.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210128286", "10002720", "EXTRAVERGINE", "SAGRA EXV 100%I BIO BOT V12x1L IT", "Ex.v. Sagra Biologico 100% ITA lt.1", 1.0, 15.00, "Vetro lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210128248", "10002747", "EXTRAVERGINE", "SAGRA EXV BOT W12x750ML CLASS IT", "Ex.v. Sagra Classico lt.0,75", 0.75, 7.50, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210121997", "10003315", "EXTRAVERGINE", "SAGRA GRAND EXV BOT W12x750ML  IT", "Ex.v. Sagra Grandulivo 0,75", 0.75, 7.50, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210127197", "10003316", "EXTRAVERGINE", "SAGRA EXV 100%I BSA BOT W12x 750ML IT", "Ex.v. Sagra Bassa Acidità 100% ITA 0,75", 0.75, 11.25, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210133792", "10003317", "EXTRAVERGINE", "SAGRA EXV 100% I BOT W 12x750ML IT", "Ex.v. Sagra 100% Italiano 0,75", 0.75, 11.25, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210131815", "10003319", "EXTRAVERGINE", "SAGRA EXV 100%I BIO BOT W12x750ML IT", "Ex.v. Sagra Biologico 100% ITA  0,75", 0.75, 11.25, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210130814", "60000444", "EXTRAVERGINE", "SAGRA EXV SPRAY C6x200ML ALLUMINIO IT", "Ex.v. Sagra Spray ml.200", 0.2, 2.00, "Spray Lt 0,20", 6, 49, 6, 294, 14, 9),
        ("8002210124387", "10003061", "EXTRAVERGINE", "SAGRA PROF EXV PET T6x2L IT", "Ex.v. Sagra Prof lt.2", 2.0, 20.00, "Pet.Lt 2", 6, 13, 4, 52, 14, 9),
        ("8002210131620", "10002724", "EXTRAVERGINE", "FBERIO EXV BOT W12x1L CLASS IT", "Ex.v. Filippo Berio Classico lt.1", 1.0, 12.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210131644", "10002725", "EXTRAVERGINE", "FBERIO EXV BOT W12x1L BSA IT", "Ex.v. Filippo Berio Bassa Acidità lt.1", 1.0, 17.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210131705", "10002726", "EXTRAVERGINE", "FBERIO EXV 100%I BOT W12x1L IT", "Ex.v. Filippo Berio 100% Italiano lt.1", 1.0, 18.00, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210131767", "10002765", "EXTRAVERGINE", "FBERIO EXV BOT W12x750ML CLASS IT", "Ex.v. Filippo Berio Classico lt.0,75", 0.75, 12.00, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210131668", "10002746", "EXTRAVERGINE", "FBERIO EXV BSA BOT W12x750ML IT", "Ex.v. Filippo Berio Bassa Acidità lt.0,75", 0.75, 17.00, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210131804", "10002768", "EXTRAVERGINE", "FBERIO EXV 100%I BOT W12x750ML IT", "Ex.v. Filippo Berio 100% Italiano lt.0,75", 0.75, 18.00, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210133013", "10003200", "EXTRAVERGINE", "FB R.O. EXV BIO 100%IT MB BOT W12X750 IT", "Ex.v. Filippo Berio Riserva Oro lt.0,75", 0.75, 19.00, "Bott.Lt 0,75", 12, 12, 5, 60, 14, 9),
        ("8002210121461", "60000544", "EXTRAVERGINE", "EX.V. BUSTINA 10mlx250 FILIPPO BERIO ITA", "Ex.v. Filippo Berio Bustina ml.10", 0.01, 0.12, "bust lt 0,01", 250, 20, 5, 100, 14, 9),
        ("8002210126572", "10003240", "OLIVA", "SAGRA OOL PUR R-PET V12X750ML CLASS IT", "Oliva Sagra RPET lt.0,75 PET", 0.75, 8.00, "Pet.Lt 0,75", 12, 12, 5, 60, 18, 12),
        ("8002210133853", "10003241", "OLIVA", "SAGRA OOL PUR R-PET V12X750ML CLASS IT", "Oliva Sagra lt.0,75", 0.75, 8.00, "Pet.Lt 0,75", 12, 12, 5, 60, 18, 12),
        ("8002210001305", "10002717", "OLIVA", "SAGRA OOL BOT W12x1L CLASS", "Oliva Sagra lt.1", 1.0, 8.00, "Bott.Lt 1", 12, 8, 5, 40, 18, 12),
        ("8002210128453", "10002718", "OLIVA", "SAGRA GRAND OOL BOT W12x1L", "Oliva Sagra Grandulivo lt.1", 1.0, 8.00, "Bott.Lt 1", 12, 8, 5, 40, 18, 12),
        ("8002210126176", "10003288", "OLIVA", "SAGRA OOL PUR R-PET T6X1.5L IT", "Oliva Sagra lt.1,5", 1.5, 12.00, "Pet.Lt 1,5", 6, 16, 4, 64, 18, 12),
        ("8002210119567", "10000537", "OLIVA", "SAGRA PROF. OOL PUR PET C2x5L IT", "Oliva Sagra Prof Lt.5", 5.0, 40.00, "Pet.Lt 5", 2, 17, 4, 68, 18, 12),
        ("8002210132436", "10002965", "OLIVA", "FBERIO OOL PUR BOT V6X500ML IT", "Oliva Filippo Berio lt.0,50", 0.5, 4.11, "Bott.Lt 0,5", 6, 30, 6, 180, 18, 12),
        ("8002210131729", "10002727", "OLIVA", "FBERIO OOL PUR BOT W12x1L IT", "Oliva Filippo Berio lt.1", 1.0, 7.75, "Bott.Lt 1", 12, 8, 5, 40, 18, 12),
        ("8002210131781", "10002766", "OLIVA", "FBERIO OOL PUR BOT W12x750ML IT", "Oliva Filippo Berio lt.0,75", 0.75, 5.97, "Bott.Lt 0,75", 12, 12, 5, 60, 18, 12),
        ("8002210122307", "10000922", "OLIVA", "FBERIO OOL PUR LAT V8x1L IT", "Oliva Filippo Berio Latta lt.1", 1.0, 8.10, "Latta lt 1", 8, 12, 5, 60, 18, 12),
        ("8002210111486", "10003307", "SEMI", "SAGRA SEM MAIS PET V12x1L IT", "Mais Sagra lt.1", 1.0, 2.00, "Pet.Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210127067", "10003286", "SEMI", "SAGRA SEM MAIS PET T6x1.5L IT", "Mais Sagrì lt.1,5", 1.5, 3.00, "Pet.Lt 1,5", 6, 16, 4, 64, 18, 12),
        ("8002210112889", "10003089", "SEMI", "SAGRA SEM MAIS PET T6x2L IT", "Mais Sagra lt.2", 2.0, 4.00, "Pet.Lt 2", 6, 13, 4, 52, 18, 12),
        ("8002210000551", "10003311", "SEMI", "SAGRA SEM ARACHIDE PET V12x1L IT", "Arachide Sagra lt.1", 1.0, 3.00, "Pet.Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210126916", "10003284", "SEMI", "SAGRI SEM ARACHIDE PET T6x1.5L IT", "Arachide Sagrì lt.1,5", 1.5, 4.50, "Pet.Lt 1,5", 6, 16, 4, 64, 18, 12),
        ("8002210112865", "10003086", "SEMI", "SAGRA SEM ARACHIDE PET T6x2L IT", "Arachide Sagra lt.2", 2.0, 6.00, "Pet.Lt 2", 6, 13, 4, 52, 18, 12),
        ("8002210116160", "10000326", "SEMI", "SAGRA PROF SEM ARACHIDE PET C2x5L IT", "Arachide Sagra Prof. Lt.5", 5.0, 15.00, "Pet lt 5", 2, 17, 4, 68, 18, 12),
        ("8002210111905", "10003310", "SEMI", "SAGRA SEM GIRAS PET V12x1L IT", "Girasole Sagra lt.1", 1.0, 2.20, "Pet.Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210126817", "10003287", "SEMI", "SAGRI SEM GIRAS PET T6x1.5L IT", "Girasole Sagrì lt.1,5", 1.5, 3.30, "Pet.Lt 1,5", 6, 16, 4, 64, 18, 12),
        ("8002210113107", "10003087", "SEMI", "SAGRA SEM GIRAS PET T6x2L IT", "Girasole Sagra lt.2", 2.0, 4.40, "Pet.Lt 2", 6, 13, 4, 52, 18, 12),
        ("8002210115453", "10003062", "SEMI", "SAGRA PROF SEM GIRAS PET C2x5L IT", "Girasole Sagra Prof Lt.5", 5.0, 11.00, "Pet lt 5", 2, 17, 4, 68, 18, 12),
        ("8002210111295", "10002933", "SEMI", "SAGRA FRIMX SEM FRITT PET V12x1L NOP IT", "Frimax Sagra lt.1", 1.0, 2.25, "Pet Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210126893", "10003285", "SEMI", "SAGRI SEM FRITT PET T6x1.5L IT", "Frimax Sagrì lt.1,5", 1.5, 3.38, "Pet.Lt 1,5", 6, 16, 4, 64, 18, 12),
        ("8002210112940", "10003085", "SEMI", "SAGRA FRIMX SEM FRITT PET T6x2L NOP IT", "Frimax Sagra lt.2", 2.0, 4.50, "Pet Lt 2", 6, 13, 4, 52, 18, 12),
        ("8002210115484", "10002644", "SEMI", "SAGRA FRIMX SEM FRITT PET C2x5L NOP IT", "Frimax Sagra lt.5", 5.0, 11.25, "Pet Lt 5", 2, 17, 4, 68, 18, 12),
        ("8002210134140", "10003327", "SEMI", "GRAZIA SEM GIRAS LAT 1x20L IT", "Frimax Spray ml.200", 0.2, 0.45, "Spray Lt 0,20", 6, 49, 6, 294, 18, 12),
        ("8002210127401", "10003309", "SEMI", "SAGRA SEM GIRAS AO PET V12x1L IT", "Girasole Alto Oleico Sagra lt.1", 1.0, 2.80, "Pet.Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210126336", "10003063", "SEMI", "SAGRA PROF SEM GIRAS AO PET C2x5L IT", "Girasole Alto Oleico Sagra Prof lt.5", 5.0, 14.00, "Pet lt 5", 2, 17, 4, 68, 18, 12),
        ("8002210129290", "10003312", "SEMI", "SAGRA SEM VINACC PET V12x1L IT", "Vinacciolo Sagra lt.1", 1.0, 5.00, "Pet.Lt 1", 12, 12, 5, 60, 18, 12),
        ("8002210130289", "10003082", "EXTRAVERGINE", "FBERIO EXV CLASS MB BOT V6x250ML IT", "Ex.v. F.Berio Anti Rab Classico lt.0,25", 0.25, 2.50, "Vetro lt 0,25", 6, 49, 5, 245, 14, 9),
        ("8002210130210", "10003081", "EXTRAVERGINE", "FBERIO EXV 100%I MB BOT V6x250ML IT", "Ex.v. F.Berio Anti Rab 100% ITA lt.0,25", 0.25, 3.00, "Vetro lt 0,25", 6, 49, 5, 245, 14, 9),
        ("8002210130340", "10003091", "EXTRAVERGINE", "FBERIO EXV CLASS MB BOT V6x500ML IT", "Ex.v. F.Berio Anti Rab Classico lt.0,50", 0.5, 4.30, "Vetro lt 0,50", 6, 31, 5, 155, 14, 9),
        ("8002210130302", "10003079", "EXTRAVERGINE", "FBERIO EXV 100%I MB BOT V6x500ML IT", "Ex.v. F.Berio Anti Rab 100% ITA lt.0,50", 0.5, 4.80, "Vetro lt 0,50", 6, 31, 5, 155, 14, 9),
        ("8002210132573", "10003072", "EXTRAVERGINE", "FBERIO EXV BOT V6x500ML TOSC IT", "Ex.v. F.Berio Toscano lt.0,50", 0.5, 10.00, "Vetro lt 0,50", 6, 31, 5, 155, 18, 12),
        ("8002210130234", "60000591", "EXTRAVERGINE", "FBERIO EXV DRES BOT V6x250ML PEP TE IT", "Ex.v. F.Berio Peperoncino lt.0,25", 0.25, 3.50, "Vetro lt 0,25", 6, 49, 5, 245, 24, 16),
        ("8002210130791", "60000590", "ACETO", "FBERIO ACE BALS BOT V6x250ML IT", "Aceto Balsamico F.Berio lt.0,25", 0.25, 2.00, "Vetro lt 0,25", 6, 48, 6, 288, 61, 41),
        ("8002210130197", "60000589", "ACETO", "FBERIO ACE BALS BOT V6x500ML IT", "Aceto Balsamico F.Berio lt.0,50", 0.5, 2.10, "Vetro lt 0,50", 6, 31, 5, 155, 61, 41)
    ]
    
    for p in prodotti_salov:
        cursor.execute("""
        INSERT OR REPLACE INTO anagrafica_master (
            ean, codice_sap, tipo_olio, descrizione_sap, descrizione_commerciale, formato_lt, confezione,
            pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet, conservazione_mesi, shelf_life_mesi
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (p[0], p[1], p[2], p[3], p[4], p[5], p[7], p[8], p[9], p[10], p[11], p[12], p[13]))
        
        cursor.execute("INSERT OR REPLACE INTO guardrail_aziendali (ean, min_net_net_g) VALUES (?, ?)", (p[0], p[6]))
        
    gdo_structure = [
        ("SELEX GRUPPO COMMERCIALE", "ALFI"), ("SELEX GRUPPO COMMERCIALE", "DIMAR"), ("SELEX GRUPPO COMMERCIALE", "ITALBRIX"), ("SELEX GRUPPO COMMERCIALE", "RIALTO"), ("SELEX GRUPPO COMMERCIALE", "ALÌ"), ("SELEX GRUPPO COMMERCIALE", "ARCA COMMERCIALE"), ("SELEX GRUPPO COMMERCIALE", "SUPERMERCATI CADORO"), ("SELEX GRUPPO COMMERCIALE", "MAXI DÌ"), ("SELEX GRUPPO COMMERCIALE", "UNICOMM"), ("SELEX GRUPPO COMMERCIALE", "CE.DI. GROS"), ("SELEX GRUPPO COMMERCIALE", "CE.DI MARCHE"), ("SELEX GRUPPO COMMERCIALE", "GMF GRANDI MAGAZZINI FIORONI"), ("SELEX GRUPPO COMMERCIALE", "MAGAZZINI GABRIELLI"), ("SELEX GRUPPO COMMERCIALE", "L’ABBONDANZA"), ("SELEX GRUPPO COMMERCIALE", "SUPER ELITE"), ("SELEX GRUPPO COMMERCIALE", "SUPEREMME"), ("SELEX GRUPPO COMMERCIALE", "CDS"), ("SELEX GRUPPO COMMERCIALE", "MEGAMARK"),
        ("GRUPPO VÉGÉ", "AMERICAN CASH"), ("GRUPPO VÉGÉ", "APULIA DISTRIBUZIONE"), ("GRUPPO VÉGÉ", "ASTA"), ("GRUPPO VÉGÉ", "BAVA"), ("GRUPPO VÉGÉ", "BENNET"), ("GRUPPO VÉGÉ", "CAPUTO SAVERIO & FIGLI"), ("GRUPPO VÉGÉ", "CARAMICO GAETANO & C."), ("GRUPPO VÉGÉ", "CENTRODET"), ("GRUPPO VÉGÉ", "COAL"), ("GRUPPO VÉGÉ", "COLONIAL SUD"), ("GRUPPO VÉGÉ", "DETERCART LOMBARDO"), ("GRUPPO VÉGÉ", "ERREGI"), ("GRUPPO VÉGÉ", "F.LLI ARENA"), ("GRUPPO VÉGÉ", "F.LLI MORGESE"), ("GRUPPO VÉGÉ", "GAMBARDELLA"), ("GRUPPO VÉGÉ", "GARGIULO & MAIELLO"), ("GRUPPO VÉGÉ", "GDA"), ("GRUPPO VÉGÉ", "GENERAL TRADE"), ("GRUPPO VÉGÉ", "G.F.E."), ("GRUPPO VÉGÉ", "GRD"), ("GRUPPO VÉGÉ", "GROSSY"), ("GRUPPO VÉGÉ", "I.S.A."), ("GRUPPO VÉGÉ", "MARKET INGROSS"), ("GRUPPO VÉGÉ", "MIGROSS"), ("GRUPPO VÉGÉ", "MODERNA 2020"), ("GRUPPO VÉGÉ", "MULTICEDI"), ("GRUPPO VÉGÉ", "MULTICEDI MCN"), ("GRUPPO VÉGÉ", "ROSSI"), ("GRUPPO VÉGÉ", "SCELGO"), ("GRUPPO VÉGÉ", "SI.D.I. PICCOLO"), ("GRUPPO VÉGÉ", "SUPERMERCATI TOSANO CEREA"), ("GRUPPO VÉGÉ", "VEGA"),
        ("CONAD", "CONAD CENTRO NORD"), ("CONAD", "COMMERCIANTI INDIPENDENTI ASSOCIATI (CIA)"), ("CONAD", "CONAD NORD OVEST"), ("CONAD", "CONAD ADRIATICO"), ("CONAD", "PAC 2000A"),
        ("COOP ITALIA", "COOP ALLEANZA 3.0"), ("COOP ITALIA", "COOP LIGURIA"), ("COOP ITALIA", "NOVA COOP"), ("COOP ITALIA", "COOP LOMBARDIA"), ("COOP ITALIA", "UNICOOP FIRENZE"), ("COOP ITALIA", "UNICOOP ETRURIA"), ("COOP ITALIA", "COOP RENO"), ("COOP ITALIA", "COOP UNIONE AMIATINA"), ("COOP ITALIA", "SAIT COOP"),
        ("C3", "BRENDOLAN ALIMENTARI"), ("C3", "C.D. GEST"), ("C3", "COLLE VERDE"), ("C3", "D’AMBROS IPERMERCATO"), ("C3", "GROS CIDAC"), ("C3", "GRUPPO BRIÒ"), ("C3", "ITALCASH"), ("C3", "LANDO F.LLI"), ("C3", "LANZA COMMERCIO DETERGENZA"), ("C3", "LEKKERLAND ITALIA"), ("C3", "LEM MARKET"), ("C3", "PERRONE"), ("C3", "PREMIUM PRICE ITALIA"), ("C3", "RETAILPRO"), ("C3", "SCUDO"), ("C3", "SUPERMERCATI GRISI"), ("C3", "SUPERMERCATI MARTINELLI"), ("C3", "SUPERMERCATI VISOTTO"), ("C3", "TO.CAL"), ("C3", "VIVO FRIULI VENEZIA GIULIA"),
        ("AGORÀ NETWORK", "GRUPPO POLI"), ("AGORÀ NETWORK", "IPERAL SUPERMERCATI"), ("AGORÀ NETWORK", "ROSSETTO TRADE"), ("AGORÀ NETWORK", "SOGEGROSS"), ("AGORÀ NETWORK", "TIGROS"),
        ("CRAI GRUPPO", "SILDA"), ("CRAI GRUPPO", "CRAI MEDITERRANEA"), ("CRAI GRUPPO", "DISTRIBUZIONE SICILIANO"), ("CRAI GRUPPO", "CRAI TIRRENO"), ("CRAI GRUPPO", "ARCEV"), ("CRAI GRUPPO", "CODÈ CRAI OVEST"), ("CRAI GRUPPO", "F.LLI IBBA"), ("CRAI GRUPPO", "AMA CRAI EST"), ("CRAI GRUPPO", "SUPERCENTRO"),
        ("DESPAR SERVIZI", "MAIORA"), ("DESPAR SERVIZI", "ERGON"), ("DESPAR SERVIZI", "FIORINO"), ("DESPAR SERVIZI", "SCS-SUPERMERCATI CONSORZIATI SARDEGNA"), ("DESPAR SERVIZI", "CENTRO3A"),
        ("D.IT DISTRIBUZIONE ITALIANA", "CE.DI. SIGMA CAMPANIA"), ("D.IT DISTRIBUZIONE ITALIANA", "CONSORZIO EUROPA"), ("D.IT DISTRIBUZIONE ITALIANA", "LOMBARDI & C."), ("D.IT DISTRIBUZIONE ITALIANA", "REALCO"), ("D.IT DISTRIBUZIONE ITALIANA", "SAN FRANCESCO"), ("D.IT DISTRIBUZIONE ITALIANA", "SISA SICILIA"), ("D.IT DISTRIBUZIONE ITALIANA", "EUROPA COMMERCIALE"), ("D.IT DISTRIBUZIONE ITALIANA", "LE DELIZIE DEL SUD"), ("D.IT DISTRIBUZIONE ITALIANA", "VA.PA."),
        ("EUROSPIN", "SPESA INTELLIGENTE"), ("EUROSPIN", "EUROSPIN TIRRENICA"), ("EUROSPIN", "EUROSPIN LAZIO"), ("EUROSPIN", "EUROSPIN PUGLIA"), ("EUROSPIN", "EUROSPIN SICILIA"),
        ("CONSORZIO CORALIS", "ALIM GROSS"), ("CONSORZIO CORALIS", "CDC"), ("CONSORZIO CORALIS", "D.IN.AL."), ("CONSORZIO CORALIS", "DUECI"), ("CONSORZIO CORALIS", "FILICE GIOVANNI"), ("CONSORZIO CORALIS", "FILICE GROUP"), ("CONSORZIO CORALIS", "GIGANTE ALIMENTARI"), ("CONSORZIO CORALIS", "GSD"), ("CONSORZIO CORALIS", "LA PRIMA"), ("CONSORZIO CORALIS", "LOMBARDO"), ("CONSORZIO CORALIS", "MAGNONE PIÙ"), ("CONSORZIO CORALIS", "MERIDIO"), ("CONSORZIO CORALIS", "PASCAR"), ("CONSORZIO CORALIS", "PREZZEMOLO&VITALE"), ("CONSORZIO CORALIS", "TUTTODISTRIBUZIONE"), ("CONSORZIO CORALIS", "VICINO A TE"),
        ("GRUPPO FINIPER CANOVA", "IPER MONTEBELLO"), ("GRUPPO FINIPER CANOVA", "UNES"),
        ("ESSELUNGA GRUPPO", "ESSELUNGA"),
        ("PAM GRUPPO", "PAM"),
        ("SELEX GRUPPO", "SELEX ")
    ]
    cursor.executemany("INSERT INTO struttura_gdo (gruppo_macro, associato_insegna) VALUES (?, ?)", gdo_structure)
    
    demo_insegne = ['COOP ALLEANZA 3.0', 'CONAD ADRIATICO', 'ESSELUNGA', 'SELEX ', 'PAM', 'CRAI TIRRENO']
    for ins in demo_insegne:
        cursor.execute("UPDATE struttura_gdo SET attivo=1 WHERE associato_insegna=?", (ins,))
        
    fallback_data = [
        # COOP ITALIA
        # Gruppo definisce S1=20% e S2=30%
        ('COOP ITALIA', '', '', 'GRUPPO', '', None, 20.0, 30.0, None, None, None, None, None, None, 1.5, 1.0, 14.0, 8.0, None, None, None, None),
        # Referenze integrano con S3 e S4 accumulandosi a S1 e S2 del Gruppo
        ('COOP ITALIA', 'COOP ITALIA SOTTOGRUPPO', '', 'REFERENZA', '8002210131620', 66.00, None, None, 12.0, 5.0, None, None, None, None, None, None, None, None, None, None, None, None),
        ('COOP ITALIA', 'COOP ITALIA SOTTOGRUPPO', '', 'REFERENZA', '8002210111110', 60.80, None, None, 15.0, 0.0, None, None, None, None, None, None, None, None, None, None, None, None),
        ('COOP ITALIA', 'COOP ITALIA SOTTOGRUPPO', '', 'REFERENZA', '8002210001305', 43.20, None, None, 12.0, 0.0, None, None, None, None, None, None, None, None, None, None, None, None),

        # ESSELUNGA GRUPPO
        # Gruppo definisce S1=35% e S2=15%
        ('ESSELUNGA GRUPPO', '', '', 'GRUPPO', '', None, 35.0, 15.0, None, None, None, None, None, None, 1.2, 1.0, 12.0, 5.0, None, None, None, None),
        # Referenze integrano con S3 e S4 accumulandosi a S1 e S2 del Gruppo
        ('ESSELUNGA GRUPPO', 'ESSELUNGA SOTTOGRUPPO', '', 'REFERENZA', '8002210131620', 40.00, None, None, 10.0, 7.0, None, None, None, None, None, None, None, None, None, None, None, None),
        ('ESSELUNGA GRUPPO', 'ESSELUNGA SOTTOGRUPPO', '', 'REFERENZA', '8002210111110', 38.00, None, None, 55.0, 0.0, None, None, None, None, None, None, None, None, None, None, None, None),
        ('ESSELUNGA GRUPPO', 'ESSELUNGA SOTTOGRUPPO', '', 'REFERENZA', '8002210001305', 24.00, None, None, 13.0, 0.0, None, None, None, None, None, None, None, None, None, None, None, None),

        # === INSERIMENTO DATI COMPLETI GRUPPO CONAD DA PDF ===
        # Regola generale di Gruppo CONAD
        ('CONAD', '', '', 'GRUPPO', '', None, 17.0, 18.0, None, None, None, None, None, None, 1.5, 1.0, 9.0, 11.0, None, None, None, None),
        
        # Referenze di Sottogruppo CONAD SOTTOGRUPPO
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210111110', 48.80, 33.0, 21.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210133440', 36.60, 33.0, 21.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210115088', 48.70, 33.0, 21.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210127562', 48.80, 33.0, 21.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210119543', 238.50, 33.0, 21.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210112827', 55.60, 33.0, 21.0, 22.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210127425', 53.70, 33.0, 21.0, 22.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210128286', 60.40, 33.0, 21.0, 22.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210128248', 45.20, 33.0, 21.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210127197', 52.30, 33.0, 21.0, 22.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210133792', 50.90, 33.0, 21.0, 22.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210130814', 30.60, 33.0, 21.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210131620', 51.70, 33.0, 21.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210131644', 59.50, 33.0, 21.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210131705', 57.90, 33.0, 21.0, 22.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210131767', 40.20, 33.0, 21.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210131668', 46.60, 33.0, 21.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210131804', 45.20, 33.0, 21.0, 22.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210133013', 57.40, 33.0, 21.0, 22.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210126572', 34.80, 33.0, 18.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210133853', 36.20, 33.0, 18.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210001305', 46.30, 33.0, 18.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210128453', 46.30, 33.0, 18.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210126176', 67.70, 33.0, 18.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210119567', 224.00, 33.0, 18.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210132436', 26.30, 33.0, 18.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210131729', 46.70, 33.0, 18.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210131781', 36.40, 33.0, 18.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210122307', 49.90, 33.0, 18.0, 24.0, 20.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210111486', 12.95, 44.0, 24.0, 33.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210127067', 19.50, 44.0, 26.0, 33.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210112889', 25.80, 44.0, 26.0, 33.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210000551', 14.15, 44.0, 24.0, 14.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210126916', 21.15, 44.0, 26.0, 14.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210112865', 28.20, 44.0, 26.0, 14.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210116160', 70.50, 44.0, 26.0, 14.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210111905', 12.70, 44.0, 24.0, 40.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210126817', 18.80, 44.0, 26.0, 40.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210113107', 25.00, 44.0, 26.0, 40.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210115453', 62.50, 44.0, 26.0, 40.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210111295', 13.10, 44.0, 24.0, 40.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210126893', 19.25, 44.0, 26.0, 40.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210112940', 25.60, 44.0, 26.0, 40.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210115484', 64.00, 44.0, 26.0, 40.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210134140', 20.00, 44.0, 26.0, 40.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210127401', 20.00, 44.0, 24.0, 40.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210126336', 95.00, 44.0, 26.0, 40.0, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210129290', 25.30, 44.0, 24.0, None, 25.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210130289', 21.50, 33.0, 21.0, 24.0, 30.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210130210', 21.80, 33.0, 21.0, 22.0, 30.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210130340', 32.20, 33.0, 21.0, 24.0, 30.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210130302', 34.30, 33.0, 21.0, 22.0, 30.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210132573', 67.00, 33.0, 21.0, 22.0, 30.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210130234', 22.70, 33.0, 21.0, None, 30.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210130791', 9.50, 33.0, 21.0, None, 30.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),
        ('CONAD', 'CONAD SOTTOGRUPPO', '', 'REFERENZA', '8002210130197', 11.80, 33.0, 21.0, None, 30.0, None, None, None, None, 1.0, 0.5, 3.00, 2.00, None, None, None, None),

        # SELEX GRUPPO
        # Gruppo definisce S1=17% e S2=18%
        ('SELEX GRUPPO', '', '', 'GRUPPO', '', None, 17.0, 18.0, None, None, None, None, None, None, 1.5, 1.0, 9.0, 11.0, None, None, None, None),
        # Referenze integrano con S3 e S4 accumulandosi a S1 e S2 del Gruppo
        ('SELEX GRUPPO', 'SELEX SOTTOGRUPPO', '', 'REFERENZA', '8002210131620', 50.00, None, None, 12.0, 9.0, None, None, None, None, None, None, None, None, None, None, None, None),
        ('SELEX GRUPPO', 'SELEX SOTTOGRUPPO', '', 'REFERENZA', '8002210111110', 44.00, None, None, 11.0, 4.0, None, None, None, None, None, None, None, None, None, None, None, None),
        ('SELEX GRUPPO', 'SELEX SOTTOGRUPPO', '', 'REFERENZA', '8002210001305', 30.00, None, None, 10.0, 4.0, None, None, None, None, None, None, None, None, None, None, None, None),
        
        # PAM GRUPPO
        # Gruppo definisce S1=15% e S2=20%
        ('PAM GRUPPO', '', '', 'GRUPPO', '', None, 15.0, 20.0, None, None, None, None, None, None, 1.4, 1.0, 11.0, 6.0, None, None, None, None),
        # Referenze integrano con S3 e S4 accumulandosi a S1 e S2 del Gruppo
        ('PAM GRUPPO', 'PAM SOTTOGRUPPO', '', 'REFERENZA', '8002210131620', 52.00, None, None, 14.0, 6.0, None, None, None, None, None, None, None, None, None, None, None, None),
        ('PAM GRUPPO', 'PAM SOTTOGRUPPO', '', 'REFERENZA', '8002210111110', 48.00, None, None, 13.0, 3.0, None, None, None, None, None, None, None, None, None, None, None, None),
        ('PAM GRUPPO', 'PAM SOTTOGRUPPO', '', 'REFERENZA', '8002210001305', 32.00, None, None, 9.0, 3.0, None, None, None, None, None, None, None, None, None, None, None, None),

        # CRAI GRUPPO
        # Gruppo definisce S1=12% e S2=25%
        ('CRAI GRUPPO', '', '', 'GRUPPO', '', None, 12.0, 25.0, None, None, None, None, None, None, 2.0, 1.0, 7.0, 12.0, None, None, None, None),
        # Referenze integrano con S3 e S4 accumulandosi a S1 e S2 del Gruppo
        ('CRAI GRUPPO', 'CRAI SOTTOGRUPPO', '', 'REFERENZA', '8002210131620', 56.00, None, None, 15.0, 8.0, None, None, None, None, None, None, None, None, None, None, None, None),
        ('CRAI GRUPPO', 'CRAI SOTTOGRUPPO', '', 'REFERENZA', '8002210111110', 50.00, None, None, 12.0, 5.0, None, None, None, None, None, None, None, None, None, None, None, None),
        ('CRAI GRUPPO', 'CRAI SOTTOGRUPPO', '', 'REFERENZA', '8002210001305', 35.00, None, None, 11.0, 5.0, None, None, None, None, None, None, None, None, None, None, None, None)
    ]
    
    cursor.executemany("""
    INSERT OR REPLACE INTO accordi_commerciali (
        gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello, listino_r,
        sconto_1, sconto_2, sconto_3, sconto_4, sconto_5,
        sconto_6, sconto_7, sconto_y, sconto_carico, sconto_pagamento,
        voce_contratto_1, voce_contratto_2, voce_contratto_3, voce_contratto_4, voce_contratto_5, note_locali
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, fallback_data)
    conn.commit()

init_db()

# ==========================================
# GESTIONE NAVIGAZIONE MENU (RIORGANIZZATA)
# ==========================================
if "main_menu_radio" not in st.session_state:
    st.session_state.main_menu_radio = "Simulatore Offerte"

if "go_to_menu" in st.session_state:
    st.session_state.main_menu_radio = st.session_state.go_to_menu
    del st.session_state.go_to_menu

st.sidebar.markdown("## Menu Principale")

menu_options = [
    
    "Simulatore Offerte", 
    "Rinnovi Contrattuali (N vs N+1)",
    "Storico Promozioni", 
    
    "Dati Anagrafici (Logistica)", 
    "Back-Office (Contratti Nazionali)", 
    "Anagrafica GDO (Clienti)",
    "Accordi Locali (Promo)",
    
    "Report Sintetico", 
    "Guida Operativa"
]

menu_selection = st.sidebar.radio("", menu_options, label_visibility="collapsed", key="main_menu_radio")

if menu_selection.startswith("---"):
    st.sidebar.warning("Seleziona una voce valida dal menu ⬆️")
    st.stop()
    
menu = menu_selection

# --- DANGER ZONE NELLA SIDEBAR ---
st.sidebar.markdown("<br><br>", unsafe_allow_html=True)
with st.sidebar.container(border=True):
    st.markdown("<h4 style='color: #991B1B; font-size: 1.1rem;'>⚠️ Danger Zone</h4>", unsafe_allow_html=True)
    if PRODUCTION_MODE:
        st.info("Ripristino disattivato in Prod.")
    else:
        st.markdown("<span style='font-size: 0.85rem;'>Ripristina il DB allo stato iniziale.</span>", unsafe_allow_html=True)
        pin_conferma = st.text_input("Digita 'RESET':", key="reset_pin_sidebar")
        if st.button("HARD RESET DB", disabled=(pin_conferma != "RESET"), use_container_width=True):
            try:
                conn_reset = sqlite3.connect(DB_FILE)
                seed_baseline_data(conn_reset)
                conn_reset.close()
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.sidebar.success("DB ripristinato!")
                st.rerun()
            except Exception as ex:
                st.sidebar.error(f"Errore: {ex}")

# --- SYSTEM METADATA FOOTER (OTTIMIZZAZIONE UX REPLICATA DA REACT) ---
st.sidebar.markdown("<br>", unsafe_allow_html=True)
st.sidebar.markdown("""
<div style="background-color: #E9E9E1; border: 1px solid #E2E2D8; padding: 12px; border-radius: 12px; font-family: 'JetBrains Mono', monospace; font-size: 0.7rem; color: #7A7E72; line-height: 1.5;">
    <div style="font-weight: bold; color: #2D3227; margin-bottom: 4px; font-family: 'Space Grotesk', sans-serif; font-size: 0.8rem;">SYSTEM METADATA</div>
    DB_RELATION: MEM_LOCAL<br/>
    ENGINE: GEOMETRIC_7_CASCADE<br/>
    STATUS: <span style="color: #5A6340; font-weight: bold;">HEALTHY ✓</span>
</div>
""", unsafe_allow_html=True)
# ---------------------------------------------

# ==========================================
# SCHEDA 1: SIMULATORE OFFERTE (Singola SKU)
# ==========================================
if menu == "Simulatore Offerte":
    conn = sqlite3.connect(DB_FILE)
    st.markdown("## Commerciale Salov - Simulatore")
    
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT gruppo_macro FROM struttura_gdo WHERE attivo=1 ORDER BY gruppo_macro")
    gruppi = [r[0] for r in cursor.fetchall()]
    
    if not gruppi:
        st.warning("Nessun cliente attivo. Vai in 'Anagrafica GDO' per attivare le insegne.")
        st.stop()

    # --- CALLBACKS DI AUTOCOMPILAZIONE REATTIVA ---
    def on_insegna_change():
        ins = st.session_state.widget_insegna
        if ins and ins != "":
            conn_cb = sqlite3.connect(DB_FILE)
            cursor_cb = conn_cb.cursor()
            cursor_cb.execute("SELECT gruppo_macro, sottogruppo FROM struttura_gdo WHERE associato_insegna=? LIMIT 1", (ins,))
            res = cursor_cb.fetchone()
            conn_cb.close()
            if res:
                st.session_state.widget_gruppo = res[0]
                st.session_state.widget_sottogruppo = res[1] or ""

    def on_gruppo_change():
        grp = st.session_state.widget_gruppo
        ins = st.session_state.widget_insegna
        if ins and ins != "":
            conn_cb = sqlite3.connect(DB_FILE)
            cursor_cb = conn_cb.cursor()
            cursor_cb.execute("SELECT COUNT(*) FROM struttura_gdo WHERE associato_insegna=? AND gruppo_macro=?", (ins, grp))
            belongs = cursor_cb.fetchone()[0] > 0
            conn_cb.close()
            if not belongs:
                st.session_state.widget_insegna = ""
                st.session_state.widget_sottogruppo = ""

    if 'widget_insegna' not in st.session_state:
        st.session_state.widget_insegna = ""
    if 'widget_gruppo' not in st.session_state:
        st.session_state.widget_gruppo = gruppi[0]
    if 'widget_sottogruppo' not in st.session_state:
        st.session_state.widget_sottogruppo = ""

    cursor.execute("SELECT DISTINCT associato_insegna FROM struttura_gdo WHERE attivo=1 ORDER BY associato_insegna")
    tutte_insegne_attive = [r[0] for r in cursor.fetchall() if r[0]]

    with st.container(border=True):
        st.markdown("#### Contesto Negoziale - Inserire Insegna")
        col_ctx1, col_ctx2, col_ctx3, col_ctx4 = st.columns(4)
        
        with col_ctx1:
            associato_sel = st.selectbox(
                "1. Insegna Locale", 
                [""] + tutte_insegne_attive, 
                key="widget_insegna", 
                on_change=on_insegna_change
            )
            
        with col_ctx2:
            gruppo_sel = st.selectbox(
                "2. Gruppo GDO", 
                gruppi, 
                key="widget_gruppo", 
                on_change=on_gruppo_change
            )
        
        cursor.execute("""
            SELECT DISTINCT sottogruppo FROM accordi_commerciali WHERE gruppo_macro=? AND sottogruppo != ''
            UNION
            SELECT DISTINCT sottogruppo FROM struttura_gdo WHERE gruppo_macro=? AND sottogruppo != '' AND sottogruppo IS NOT NULL
            ORDER BY sottogruppo
        """, (gruppo_sel, gruppo_sel))
        sottogruppi = [r[0] for r in cursor.fetchall()]
        if not sottogruppi: 
            sottogruppi = [""]
            
        with col_ctx3:
            sottogruppo_sel = st.selectbox(
                "3. Sottogruppo GDO", 
                sottogruppi, 
                key="widget_sottogruppo"
            )

        cursor.execute("""
            SELECT a.ean, a.descrizione_commerciale, a.tipo_olio, COALESCE(g.min_net_net_g, 0.0), a.codice_sap, a.formato_lt,
                   a.pezzi_cartone, a.cartoni_strato, a.strati_pallet, a.cartoni_pallet
            FROM anagrafica_master a
            LEFT JOIN guardrail_aziendali g ON a.ean = g.ean
        """)
        prodotti = cursor.fetchall()
        prodotti_dict = {f"{p[1]} [EAN: {p[0]}]": (p[0], p[2], p[3], p[4], p[5], p[6], p[7], p[8], p[9]) for p in prodotti}
        
        if 'clone_ean_pending' in st.session_state:
            ean_to_find = st.session_state.pop('clone_ean_pending')
            for p_key in prodotti_dict.keys():
                if ean_to_find in p_key:
                    st.session_state['widget_prodotto'] = p_key
                    break
                    
        with col_ctx4:
            prodotto_scelto = st.selectbox("4. Referenza Salov", list(prodotti_dict.keys()), key="widget_prodotto")

    ean, tipo_olio, min_net_net_g, codice_sap, formato_lt, pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet = prodotti_dict[prodotto_scelto]
    
    # Merge strutturale e locale
    contract = get_merged_contract(conn, gruppo_sel, sottogruppo_sel, associato_sel, ean, tipo_olio)

    # --- AVVISO ACCORDI LOCALI (OTTIMIZZAZIONE UX REPLICATA DA REACT) ---
    cursor.execute("SELECT COUNT(*) FROM accordi_commerciali WHERE gruppo_macro=? AND associato_insegna=? AND associato_insegna != '' AND chiave_livello=?", (gruppo_sel, associato_sel, ean))
    has_local = cursor.fetchone()[0] > 0
    if has_local:
        st.markdown(f"""
        <div class="alert-box alert-success" style="margin-top: 15px;">
            <strong>Accordi Locali Attivi sul Territorio:</strong> Sono configurate promozioni dirette per l'insegna <strong>{associato_sel}</strong>.<br/>
            Sconto 6: <strong>{contract.sconto_6 or 0}%</strong> | Sconto 7: <strong>{contract.sconto_7 or 0}%</strong> | Sconto Y: <strong>{contract.sconto_y or 0}%</strong> ereditati nel simulatore.
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div class="alert-box alert-info" style="margin-top: 15px;">
            <strong>Nessun Accordo Locale Attivo:</strong> Non sono presenti accordi promozionali territoriali registrati per l'insegna <strong>{associato_sel}</strong>. Verranno applicate esclusivamente le condizioni nazionali del Sottogruppo.
        </div>
        """, unsafe_allow_html=True)
    # -----------------------------

    if contract.listino_r is None:
        st.error("ATTENZIONE: PRODOTTO FUORI ASSORTIMENTO PER QUESTO CLIENTE")
        st.stop()

    col_m1, col_m2, col_m3 = st.columns(3)
    col_m1.metric("Listino Base (R)", fmt_it(float(contract.listino_r), is_euro=True))
    col_m2.metric("Soglia Minima Net Net (G)", fmt_it(float(min_net_net_g), is_euro=True))
    
    # Renderizzazione Micro-Badge ereditarietà coerente con React
    with col_m3:
        st.markdown(f"""
        <div style="background-color: #FFFFFF; padding: 20px; border-radius: 16px; border: 1px solid #E2E2D8; box-shadow: 0 1px 3px rgba(45,50,39,0.05); display: flex; flex-direction: column; justify-content: center; height: 100%;">
            <div style="font-size: 0.75rem; color: #7A7E72; text-transform: uppercase; letter-spacing: 0.1em; font-weight: 700; margin-bottom: 8px;">LIVELLO ERIDITARIETÀ RISOLTO</div>
            <div>{render_badge(contract.livello_risolto)}</div>
        </div>
        """, unsafe_allow_html=True)

    st.divider()

    col_met1, col_met2 = st.columns([1, 1])
    with col_met1:
        st.markdown("#### Metodologia di Calcolo")
        metodo_lavoro = st.radio(
            "Seleziona l'approccio negoziale:",
            ["A. Partenza da Prezzo Target (Calcolo automatico Sconto Promo)", "B. Tentativi Spot Manuali (Immissione Sconto Promo libera)"],
            horizontal=False,
            label_visibility="collapsed",
            key="widget_metodo"
        )

    with col_met2:
        if "A. Partenza" in metodo_lavoro:
            st.markdown("#### Obiettivo Economico")
            target_net_net = st.number_input(
                "PREZZO TARGET NET NET DESIDERATO (Euro/Pz)", 
                min_value=0.0, 
                value=float(min_net_net_g), 
                step=0.10
            )
        else:
            target_net_net = 0.0

    st.markdown("<br>", unsafe_allow_html=True)

    col_l1, col_l2 = st.columns(2)
    
    with col_l1:
        st.markdown("#### Scontistiche Utilizzabili")
        with st.container(border=True):
            sconto_y = st.number_input("Sconto Continuativo Y (%)", min_value=0.0, max_value=100.0, value=float(contract.sconto_y or 0.0), step=0.5)
            if contract.sconto_y and float(contract.sconto_y) > 0:
                st.markdown(f"<div class='alert-box alert-warning'>ATTENZIONE - SE LO SCONTO CONTINUATIVO DERIVA DA UN ACCORDO LOCALE NON LO SI PUO' VARIARE SENZA UN NUOVO ACCORDO - valore attuale: {fmt_it(float(contract.sconto_y), is_pct=True)}</div>", unsafe_allow_html=True)
            
            if "A. Partenza" in metodo_lavoro:
                st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
                st.markdown("<h5 style='color: #1A3E2F; margin-bottom: 5px;'>Leva Promozionale Diretta</h5>", unsafe_allow_html=True)
                sconto_aa = st.number_input("Sconto Unitario in fattura (Euro/Pz) [AA]", min_value=0.0, step=0.05, key="widget_aa")
                sconto_z_input = 0.0
            else:
                st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
                st.markdown("**Leve Promozionali**")
                sconto_z_input = st.number_input("Sconto Promozionale (%) [Z] (Manuale)", min_value=0.0, max_value=100.0, step=0.5, key="widget_z")
                sconto_aa = st.number_input("Sconto Unitario in fattura (Euro/Pz) [AA]", min_value=0.0, step=0.05, key="widget_aa")
                
            sconto_z = safe_dec(sconto_z_input)

    # Input protetto per il PricingEngine
    engine_input = PricingInput(
        listino_r=safe_dec(contract.listino_r),
        sconto_1=safe_dec(contract.sconto_1),
        sconto_2=safe_dec(contract.sconto_2),
        sconto_3=safe_dec(contract.sconto_3),
        sconto_4=safe_dec(contract.sconto_4),
        sconto_5=safe_dec(contract.sconto_5),
        sconto_6=safe_dec(contract.sconto_6),
        sconto_7=safe_dec(contract.sconto_7),
        sconto_y=safe_dec(sconto_y),
        sconto_z=sconto_z,
        sconto_aa=safe_dec(sconto_aa),
        sconto_carico=safe_dec(contract.sconto_carico),
        sconto_pagamento=safe_dec(contract.sconto_pagamento),
        voce_i=safe_dec(contract.voce_i),
        voce_ii=safe_dec(contract.voce_ii),
        voce_iii=safe_dec(contract.voce_iii),
        voce_iv=safe_dec(contract.voce_iv),
        voce_v=safe_dec(contract.voce_v),
        min_net_net_g=safe_dec(min_net_net_g)
    )

    if "A. Partenza" in metodo_lavoro:
        target_dec = safe_dec(target_net_net)
        sconto_z = PricingEngine.calculate_inverse(target_dec, engine_input, "Z")
        engine_input = replace(engine_input, sconto_z=sconto_z)

    with col_l2:
        st.markdown("#### Limiti Promozionali per il net net minimo - valori di riferimento - nel caso A è lo sconto applicato")
        with st.container(border=True):
            if "A. Partenza" in metodo_lavoro:
                engine_max_z = replace(engine_input, sconto_z=Decimal("0.00"))
                z_max_consentito = PricingEngine.calculate_inverse(safe_dec(min_net_net_g), engine_max_z, "Z")
                st.number_input("Sconto Promo MAX Consentito [Z]", value=float(z_max_consentito), disabled=True, format="%.2f")
                st.markdown("<div style='height: 85px;'></div>", unsafe_allow_html=True) 
            else:
                engine_max_z = replace(engine_input, sconto_z=Decimal("0.00"))
                z_max_consentito = PricingEngine.calculate_inverse(safe_dec(min_net_net_g), engine_max_z, "Z")
                st.number_input("Sconto Promo MAX Consentito [Z]", value=float(z_max_consentito), disabled=True, format="%.2f")
                
                st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
                
                engine_max_aa = replace(engine_input, sconto_aa=Decimal("0.00"))
                aa_max_consentito = PricingEngine.calculate_inverse(safe_dec(min_net_net_g), engine_max_aa, "AA")
                st.number_input("Sconto Unitario MAX Consentito [AA]", value=float(aa_max_consentito), disabled=True, format="%.2f")

    result = PricingEngine.calculate(engine_input)

    st.divider()

    st.markdown("#### Risultato Simulazione")
    res_col1, res_col2 = st.columns([2, 1])
    
    with res_col1:
        with st.expander("Verifica Margine e Stato (Contrattuale)", expanded=True):
            st.metric("PREZZO NET NET RISULTANTE (AM)", fmt_it(float(result.net_net_finale), 3, is_euro=True))
            st.metric("SOGLIA MINIMA NET NET (G)", fmt_it(float(min_net_net_g), 3, is_euro=True))
            if result.guardrail_ok:
                st.success(f"APPROVATO - Margine sicuro. Delta: +{fmt_it(float(result.delta_vs_min), 3, is_euro=True)}")
            else:
                st.error(f"BLOCCATO! - Sotto soglia di {fmt_it(float(abs(result.delta_vs_min)), 3, is_euro=True)}")
            
    with res_col2:
        with st.expander("Finestra Temporale Promo", expanded=True):
            col_d1, col_d2 = st.columns(2)
            with col_d1:
                sell_in_dal = st.date_input("Inizio Sell-In", date.today())
                sell_out_dal = st.date_input("Inizio Sell-Out", date.today())
            with col_d2:
                sell_in_al = st.date_input("Fine Sell-In", date.today())
                sell_out_al = st.date_input("Fine Sell-Out", date.today())

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown("#### Contributi Promozionali Extra (Sell-Out)")
    with st.container(border=True):
        col_v1, col_v2, col_v3 = st.columns(3)
        with col_v1:
            volumi_stimati = st.number_input("Volumi Stimati (Pezzi)", min_value=0, step=100, key="widget_volumi")
        with col_v2:
            contributo_fisso = st.number_input("Contributo Fisso Totale (€)", min_value=0.0, step=50.0, key="widget_fisso")
        with col_v3:
            contributo_pezzo = st.number_input("Contributo a Pezzo (€/Pz)", min_value=0.0, step=0.05, key="widget_pezzo")

        costo_totale_extra = contributo_fisso + (contributo_pezzo * volumi_stimati)
        impatto_unitario_extra = Decimal("0.00")
        net_net_post_promo = result.net_net_finale
        mostra_impatto = False
        
        if (volumi_stimati > 0 and contributo_fisso > 0) or (contributo_pezzo > 0):
            mostra_impatto = True
            if volumi_stimati > 0:
                impatto_unitario_extra = Decimal(str(contributo_pezzo)) + (Decimal(str(contributo_fisso)) / Decimal(str(volumi_stimati)))
            else:
                impatto_unitario_extra = Decimal(str(contributo_pezzo))
            
            net_net_post_promo = result.net_net_finale - impatto_unitario_extra
            
            st.markdown(f"<div class='alert-box alert-warning'><strong>Costo Extra Totale:</strong> {fmt_it(float(costo_totale_extra), is_euro=True)} | <strong>Impatto Unitario:</strong> -{fmt_it(float(impatto_unitario_extra), 3, is_euro=True)}/Pz ➔ <strong>NET NET POST-VOLANTINO: {fmt_it(float(net_net_post_promo), 3, is_euro=True)}</strong></div>", unsafe_allow_html=True)
        elif costo_totale_extra > 0:
            st.markdown(f"<div class='alert-box alert-info'><strong>Costo Extra Totale:</strong> {fmt_it(float(costo_totale_extra), is_euro=True)} (Volumi non inseriti, impatto unitario non calcolabile)</div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown("#### Struttura di Costo (Waterfall)")
    waterfall_data = [{"Fase Pricing": step.fase, "Valore Unitario": fmt_it(float(step.valore), 3, is_euro=True), "Dettaglio Operazione": step.descrizione} for step in result.steps]
    if mostra_impatto:
        waterfall_data.append({"Fase Pricing": "Impatto Extra (Sell-Out)", "Valore Unitario": fmt_it(float(net_net_post_promo), 3, is_euro=True), "Dettaglio Operazione": f"-{fmt_it(float(impatto_unitario_extra), 3, is_euro=True)}/Pz"})
        
    st.dataframe(pd.DataFrame(waterfall_data), use_container_width=True, hide_index=True)

    # --- GRAFICO A CASCATA (WATERFALL) ---
    st.markdown("#### 📊 Grafico a Cascata (Impatto Visivo)")
    x_vals = []
    y_vals = []
    measures = []
    text_vals = []
    
    prev_val = 0.0
    for i, step in enumerate(result.steps):
        x_vals.append(step.fase)
        current_val = float(step.valore)
        if i == 0:
            measures.append("relative")
            y_vals.append(current_val)
            text_vals.append(fmt_it(current_val, 3))
        elif i == len(result.steps) - 1:
            measures.append("total")
            y_vals.append(current_val)
            text_vals.append(fmt_it(current_val, 3))
        else:
            measures.append("relative")
            delta = current_val - prev_val
            y_vals.append(delta)
            text_vals.append(fmt_it(delta, 3))
        prev_val = current_val
        
    if mostra_impatto:
        measures[-1] = "relative"
        y_vals[-1] = current_val - prev_val if len(result.steps) > 1 else current_val
        
        x_vals.append("Impatto Extra")
        measures.append("relative")
        y_vals.append(-float(impatto_unitario_extra))
        text_vals.append("-" + fmt_it(float(impatto_unitario_extra), 3))
        
        x_vals.append("Net Net Post-Promo")
        measures.append("total")
        y_vals.append(float(net_net_post_promo))
        text_vals.append(fmt_it(float(net_net_post_promo), 3))
        
    fig = go.Figure(go.Waterfall(
        orientation="v", measure=measures, x=x_vals, y=y_vals, text=text_vals, textposition="outside",
        decreasing={"marker":{"color":"#A34A3F"}}, increasing={"marker":{"color":"#8A9A5B"}}, totals={"marker":{"color":"#5A6340"}}
    ))
    
    fig.update_layout(
        title="Evoluzione del Margine Unitario (€)", 
        waterfallgap=0.2, 
        margin=dict(t=40, b=40, l=40, r=40), 
        showlegend=False,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(family="Space Grotesk, sans-serif", size=11, color="#2D3227")
    )
    st.plotly_chart(fig, use_container_width=True)
    # ------------------------------------

    st.divider()
    
    col_act1, col_act2 = st.columns(2)
    
    with col_act1:
        st.markdown("#### Salva nel CRM")
        with st.container(border=True):
            st.markdown("#### Salva nel CRM")
            stato_promo = st.radio("Stato Promozione", ["Proposta", "Confermata"], horizontal=True)
            note_promo = st.text_input("Note (es. Volantino)")
            if st.button("Salva Promozione", type="primary", use_container_width=True):
                try:
                    c_save = conn.cursor()
                    c_save.execute("""
                        INSERT INTO storico_promo (
                            stato_promo, gruppo_macro, sottogruppo, associato_insegna, ean, descrizione_commerciale, listino_r, sconto_y, sconto_z, sconto_aa,
                            net_net_am, volumi_stimati, contributo_fisso, contributo_pezzo, costo_totale_extra, note,
                            sell_in_dal, sell_in_al, sell_out_dal, sell_out_al, min_net_net_g, net_net_post_promo
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        stato_promo, gruppo_sel, sottogruppo_sel, associato_sel, ean, prodotto_scelto.split(" [EAN:")[0],
                        float(contract.listino_r), float(sconto_y), float(sconto_z), float(sconto_aa), float(result.net_net_finale),
                        volumi_stimati, contributo_fisso, contributo_pezzo, costo_totale_extra, note_promo,
                        sell_in_dal.strftime('%Y-%m-%d'), sell_in_al.strftime('%Y-%m-%d'), sell_out_dal.strftime('%Y-%m-%d'), sell_out_al.strftime('%Y-%m-%d'),
                        float(min_net_net_g), float(net_net_post_promo)
                    ))
                    conn.commit()
                    st.success("Salvato!")
                except Exception as e:
                    st.error(f"Errore: {e}")

    with col_act2:
        st.markdown("#### Esporta Proposta")
        with st.container(border=True):
            st.markdown("Genera il file Excel ufficiale con i dettagli della simulazione corrente.")
            st.markdown("<br>", unsafe_allow_html=True)
            
            def genera_scheda_negoziale():
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Proposta_Commerciale"
                ws.views.sheetView[0].showGridLines = True
                
                font_title = Font(name="Arial", size=15, bold=True, color="FFFFFF")
                font_section = Font(name="Arial", size=11, bold=True, color="000000")
                font_label = Font(name="Arial", size=10, bold=True)
                font_value = Font(name="Arial", size=10)
                fill_header = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
                fill_sub = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                thin_border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'), top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
                
                ws.merge_cells('A1:D1')
                ws['A1'] = "SALOV S.p.A. - SCHEDA PROPOSTA COMMERCIALE"
                ws['A1'].font = font_title
                ws['A1'].fill = fill_header
                ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 40
                
                ws['A3'] = "ANAGRAFICA GDO"
                ws['A3'].font = font_section
                ws['A3'].fill = fill_sub
                ws.merge_cells('A3:D3')
                
                ws['A4'] = "Gruppo Macro:"
                ws['B4'] = gruppo_sel
                ws['A5'] = "Sottogruppo:"
                ws['B5'] = sottogruppo_sel
                ws['A6'] = "Insegna Locale / Associato:"
                ws['B6'] = associato_sel
                
                for r in range(4, 7):
                    ws[f'A{r}'].font = font_label
                    ws[f'B{r}'].font = font_value
                
                ws['A8'] = "DETTAGLIO REFERENZA"
                ws['A8'].font = font_section
                ws['A8'].fill = fill_sub
                ws.merge_cells('A8:D8')
                
                ws['A9'] = "Descrizione Articolo:"
                ws['B9'] = prodotto_scelto.split(" [EAN:")[0]
                ws['A10'] = "EAN:"
                ws['B10'] = ean
                ws['A11'] = "Codice SAP:"
                ws['B11'] = codice_sap
                ws['A12'] = "Formato:"
                ws['B12'] = f"{formato_lt} Litri"
                
                for r in range(9, 13):
                    ws[f'A{r}'].font = font_label
                    ws[f'B{r}'].font = font_value
                    
                ws['A14'] = "DATI LOGISTICI E PALLETTIZZAZIONE"
                ws['A14'].font = font_section
                ws['A14'].fill = fill_sub
                ws.merge_cells('A14:D14')
                
                ws['A15'] = "Pezzi per Cartone:"
                ws['B15'] = pezzi_cartone if pezzi_cartone is not None else 0
                ws['A16'] = "Cartoni per Strato:"
                ws['B16'] = cartoni_strato if cartoni_strato is not None else 0
                ws['A17'] = "Strati per Pallet:"
                ws['B17'] = strati_pallet if strati_pallet is not None else 0
                ws['A18'] = "Cartoni per Pallet:"
                ws['B18'] = cartoni_pallet if cartoni_pallet is not None else 0
                ws['A19'] = "Pezzi Totali per Pallet:"
                ws['B19'] = (pezzi_cartone or 0) * (cartoni_pallet or 0)
                
                for r in range(15, 20):
                    ws[f'A{r}'].font = font_label
                    ws[f'B{r}'].font = font_value
                    
                ws['A21'] = "FINESTRE TEMPORALI PROMO"
                ws['A21'].font = font_section
                ws['A21'].fill = fill_sub
                ws.merge_cells('A21:D21')
                
                ws['A22'] = "Periodo Sell-In:"
                ws['B22'] = f"Dal {sell_in_dal.strftime('%d/%m/%Y')} al {sell_in_al.strftime('%d/%m/%Y')}"
                ws['A23'] = "Periodo Sell-Out:"
                ws['B23'] = f"Dal {sell_out_dal.strftime('%d/%m/%Y')} al {sell_out_al.strftime('%d/%m/%Y')}"
                
                for r in range(22, 24):
                    ws[f'A{r}'].font = font_label
                    ws[f'B{r}'].font = font_value
                
                ws['A25'] = "CASCATA DI PRICING NEGOZIALE"
                ws['A25'].font = font_section
                ws['A25'].fill = fill_sub
                ws.merge_cells('A25:D25')
                
                ws['A26'] = "Elemento di Costo"
                ws['B26'] = "Valore"
                ws['C26'] = "Tipologia Operazione"
                for col in ['A', 'B', 'C']:
                    ws[f'{col}26'].font = font_label
                    
                row_idx = 27
                for step in result.steps:
                    ws.cell(row=row_idx, column=1, value=step.fase).font = font_value
                    ws.cell(row=row_idx, column=2, value=float(step.valore)).font = font_value
                    ws.cell(row=row_idx, column=2).number_format = '#,##0.000 €'
                    ws.cell(row=row_idx, column=3, value=step.descrizione).font = font_value
                    row_idx += 1
                    
                if mostra_impatto:
                    ws.cell(row=row_idx, column=1, value="Impatto Extra (Sell-Out)").font = font_value
                    ws.cell(row=row_idx, column=2, value=float(net_net_post_promo)).font = font_value
                    ws.cell(row=row_idx, column=2).number_format = '#,##0.000 €'
                    ws.cell(row=row_idx, column=3, value=f"-{float(impatto_unitario_extra):.3f} €/Pz").font = font_value
                    row_idx += 1
                    
                ws.cell(row=row_idx+1, column=1, value="SOGLIA MINIMA AM (G):").font = font_label
                ws.cell(row=row_idx+1, column=2, value=float(min_net_net_g)).font = font_value
                ws.cell(row=row_idx+1, column=2).number_format = '#,##0.00 €'
                
                ws.cell(row=row_idx+2, column=1, value="DELTA DI MARGINE VS SOGLIA:").font = font_label
                ws.cell(row=row_idx+2, column=2, value=float(result.delta_vs_min)).font = font_value
                ws.cell(row=row_idx+2, column=2).number_format = '#,##0.00 €'
                
                ws.cell(row=row_idx+3, column=1, value="STATO DEL MARGINE:").font = font_label
                stato_txt = "VERDE (APPROVATO)" if result.guardrail_ok else "ROSSO (SOTTO SOGLIA)"
                ws.cell(row=row_idx+3, column=2, value=stato_txt).font = font_label
                
                if costo_totale_extra > 0:
                    row_idx += 5
                    ws.cell(row=row_idx, column=1, value="CONTRIBUTI EXTRA (SELL-OUT)").font = font_section
                    ws.cell(row=row_idx, column=1).fill = fill_sub
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
                    
                    ws.cell(row=row_idx+1, column=1, value="Volumi Stimati (Pz):").font = font_label
                    ws.cell(row=row_idx+1, column=2, value=volumi_stimati).font = font_value
                    
                    ws.cell(row=row_idx+2, column=1, value="Costo Totale Extra:").font = font_label
                    ws.cell(row=row_idx+2, column=2, value=costo_totale_extra).font = font_value
                    ws.cell(row=row_idx+2, column=2).number_format = '#,##0.00 €'
                    
                    if volumi_stimati > 0:
                        ws.cell(row=row_idx+3, column=1, value="Net Net Post-Volantino:").font = font_label
                        ws.cell(row=row_idx+3, column=2, value=float(net_net_post_promo)).font = font_value
                        ws.cell(row=row_idx+3, column=2).number_format = '#,##0.000 €'
                
                ws.column_dimensions['A'].width = 32
                ws.column_dimensions['B'].width = 38
                ws.column_dimensions['C'].width = 45
                
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
                    for cell in row:
                        cell.border = thin_border
                
                buffer = io.BytesIO()
                wb.save(buffer)
                return buffer.getvalue()

            st.download_button(
                label="Scarica Excel",
                data=genera_scheda_negoziale(),
                file_name=f"Proposta_{associato_sel}_{codice_sap}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    conn.close()

# ==========================================
# NUOVA SCHEDA: RINNOVI CONTRATTUALI (N vs N+1)
# ==========================================
elif menu == "Rinnovi Contrattuali (N vs N+1)":
    st.title("Simulatore per Rinnovi Contrattuali (N vs N+1)")
    st.markdown("Analisi differenziale dei margini, calcolo dello Spazio Promo e Roll-up per Sub-Categorie.")
    
    anno_corrente = date.today().year
    conn = sqlite3.connect(DB_FILE)
    
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS proposte_rinnovi (
            id INTEGER PRIMARY KEY AUTOINCREMENT, 
            data_salvataggio TIMESTAMP DEFAULT CURRENT_TIMESTAMP, 
            nome_proposta TEXT, 
            gruppo_macro TEXT, 
            sottogruppo TEXT, 
            associato_insegna TEXT, 
            global_carico REAL, 
            global_pagamento REAL, 
            dati_json TEXT
        )
    """)
    conn.commit()
    
    def to_excel_bytes(df):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Export')
        return output.getvalue()
    
    if 'global_carico' not in st.session_state:
        st.session_state.global_carico = 0.0
    if 'global_pagamento' not in st.session_state:
        st.session_state.global_pagamento = 0.0
    if 'rinnovi_gruppo' not in st.session_state:
        st.session_state.rinnovi_gruppo = "Nessuno"
    if 'rinnovi_sottogruppo' not in st.session_state:
        st.session_state.rinnovi_sottogruppo = ""

    associato_sel = ""

    st.markdown("#### 1. Contesto di Riferimento (Pre-compilazione Anno N)")
    
    with st.container(border=True):
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT gruppo_macro FROM struttura_gdo WHERE attivo=1 ORDER BY gruppo_macro")
        gruppi = [r[0] for r in cursor.fetchall()]
        
        col_selectors, col_actions = st.columns([3, 2])
        
        with col_selectors:
            st.markdown("**1. Configurazione GDO (Accordo Quadro)**")
            col_ctx1, col_ctx2 = st.columns(2)
            
            with col_ctx1:
                gruppo_sel = st.selectbox(
                    "Gruppo GDO", 
                    ["Nessuno"] + gruppi, 
                    key="rinnovi_gruppo"
                )
            
            sottogruppi = []
            if gruppo_sel != "Nessuno":
                cursor.execute("""
                    SELECT DISTINCT sottogruppo FROM accordi_commerciali WHERE gruppo_macro=? AND sottogruppo != ''
                    UNION
                    SELECT DISTINCT sottogruppo FROM struttura_gdo WHERE gruppo_macro=? AND sottogruppo != '' AND sottogruppo IS NOT NULL
                    ORDER BY sottogruppo
                """, (gruppo_sel, gruppo_sel))
                sottogruppi = [r[0] for r in cursor.fetchall()]
                
            with col_ctx2:
                sottogruppo_sel = st.selectbox(
                    "Sottogruppo GDO", 
                    [""] + sottogruppi, 
                    key="rinnovi_sottogruppo"
                )
                
        with col_actions:
            st.markdown("**2. Azioni & Salvataggi Scenario**")
            col_btn_grp1, col_btn_grp2 = st.columns(2)
            with col_btn_grp1:
                btn_carica = st.button("Carica Baseline 🔄", type="primary", use_container_width=True, help="Importa dal DB le condizioni del Sottogruppo/Gruppo.")
                btn_mock = st.button("Dati di Test (Mock) 🧪", use_container_width=True, help="Popola lo scenario corrente con dati di test simulati.")
            with col_btn_grp2:
                nome_scenario = st.text_input("Nome Proposta per Salvare", placeholder="Es. Scenario Q3", label_visibility="collapsed")
                btn_salva = st.button("Salva Scenario 💾", type="secondary", use_container_width=True, help="Salva lo stato corrente della griglia dei rinnovi per riprenderlo in futuro.")

    if btn_carica:
        if gruppo_sel != "Nessuno":
            df_temp = st.session_state.rinnovi_df.copy()
            first_contract = True
            
            def safe_float(v): return float(v) if v is not None else 0.0
            
            for idx, row in df_temp.iterrows():
                contract = get_merged_contract(conn, gruppo_sel, sottogruppo_sel, associato_sel, row['ean'], row['tipo_olio'])
                
                if contract.listino_r is None:
                    contract.listino_r = get_listino_strutturale(conn, gruppo_sel, sottogruppo_sel, row['ean'])
                
                if contract.listino_r is not None:
                    if first_contract:
                        st.session_state.global_carico = safe_float(contract.sconto_carico)
                        st.session_state.global_pagamento = safe_float(contract.sconto_pagamento)
                        first_contract = False
                        
                    listino_n = safe_float(contract.listino_r)
                    df_temp.at[idx, '[N] Listino €'] = listino_n
                    df_temp.at[idx, '[N+1] Listino €'] = listino_n
                    
                    p = Decimal(str(listino_n))
                    sconti_strutturali = [contract.sconto_1, contract.sconto_2, contract.sconto_3, contract.sconto_4, contract.sconto_5]
                    
                    for s in sconti_strutturali:
                        val_s = Decimal(str(s)) if s is not None else Decimal('0')
                        p = p * (Decimal('1') - (val_s / Decimal('100')))
                    
                    sc_fatt_eq = 0.0
                    if listino_n > 0:
                        sc_fatt_eq = float(((Decimal('1') - (p / Decimal(str(listino_n)))) * Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    
                    df_temp.at[idx, '[N] Sc. Fattura %'] = sc_fatt_eq
                    df_temp.at[idx, '[N+1] Sc. Fattura %'] = sc_fatt_eq
                    
                    pfa_tot = safe_float(contract.voce_i) + safe_float(contract.voce_ii) + safe_float(contract.voce_iii) + safe_float(contract.voce_iv) + safe_float(contract.voce_v)
                    df_temp.at[idx, '[N] Contratto %'] = pfa_tot
                    df_temp.at[idx, '[N+1] Contratto %'] = pfa_tot
                    
                    df_temp.at[idx, 'S1 %'] = safe_float(contract.sconto_1)
                    df_temp.at[idx, 'S2 %'] = safe_float(contract.sconto_2)
                    df_temp.at[idx, 'S3 %'] = safe_float(contract.sconto_3)
                    df_temp.at[idx, 'S4 %'] = safe_float(contract.sconto_4)
                    df_temp.at[idx, 'S5 %'] = safe_float(contract.sconto_5)
                    df_temp.at[idx, 'PFA I %'] = safe_float(contract.voce_i)
                    df_temp.at[idx, 'PFA II %'] = safe_float(contract.voce_ii)
                    df_temp.at[idx, 'PFA III %'] = safe_float(contract.voce_iii)
                    df_temp.at[idx, 'PFA IV %'] = safe_float(contract.voce_iv)
                    df_temp.at[idx, 'PFA V %'] = safe_float(contract.voce_v)
                    
            st.session_state.rinnovi_df = df_temp
            st.success("Condizioni contrattuali ricaricate con successo.")
            st.rerun()
        else:
            st.warning("Seleziona un Gruppo GDO valido prima di caricare i dati.")

    if btn_mock:
        df_mock = st.session_state.rinnovi_df.copy()
        st.session_state.global_carico = 2.0
        st.session_state.global_pagamento = 1.5
        for idx, row in df_mock.iterrows():
            floor = row['Minimo Net Net €'] if row['Minimo Net Net €'] > 0 else 3.0
            vol = random.randint(10, 100) * 100
            
            listino_n = float(round(floor * 1.5, 2))
            sc_fatt_n = 10.0
            pfa_n = 5.0
            
            df_mock.at[idx, '[N] Listino €'] = listino_n
            df_mock.at[idx, '[N] Sc. Fattura %'] = sc_fatt_n
            df_mock.at[idx, '[N] Contratto %'] = pfa_n
            
            df_mock.at[idx, '[N+1] Volumi'] = int(vol * 1.05)
            df_mock.at[idx, '[N+1] Listino €'] = float(round(floor * 1.6, 2))
            df_mock.at[idx, '[N+1] Sc. Fattura %'] = 12.0
            df_mock.at[idx, '[N+1] Contratto %'] = 5.0
            
            df_mock.at[idx, 'S1 %'] = 10.0
            df_mock.at[idx, 'S2 %'] = 2.22 
            df_mock.at[idx, 'PFA I %'] = 5.0
            
        st.session_state.rinnovi_df = df_mock
        st.success("Dati di test caricati con successo.")
        st.rerun()

    if btn_salva:
        if not nome_scenario.strip():
            st.warning("⚠️ Inserisci un nome per la proposta commerciale prima di procedere al salvataggio.")
        elif gruppo_sel == "Nessuno":
            st.warning("⚠️ Seleziona una configurazione GDO (Gruppo) valida prima di salvare.")
        else:
            dati_json = st.session_state.rinnovi_df.to_json(orient='records')
            cursor.execute("""
                INSERT INTO proposte_rinnovi (nome_proposta, gruppo_macro, sottogruppo, associato_insegna, global_carico, global_pagamento, dati_json)
                VALUES (?, ?, ?, '', ?, ?, ?)
            """, (nome_scenario, gruppo_sel, sottogruppo_sel, float(st.session_state.global_carico), float(st.session_state.global_pagamento), dati_json))
            conn.commit()
            st.success(f"💾 Scenario '{nome_scenario}' salvato correttamente!")
            st.rerun()

    query = """
        SELECT a.ean, a.descrizione_commerciale, a.tipo_olio, COALESCE(g.min_net_net_g, 0.0) as min_net_net_g
        FROM anagrafica_master a
        LEFT JOIN guardrail_aziendali g ON a.ean = g.ean
    """
    df_base = pd.read_sql_query(query, conn)
    df_base['Sub-Categoria'] = df_base.apply(get_subcat, axis=1)
    df_base['Categoria'] = df_base['tipo_olio']
    df_base = df_base.rename(columns={'descrizione_commerciale': 'Prodotto', 'min_net_net_g': 'Minimo Net Net €'})

    dettaglio_cols = [
        'S1 %', 'S2 %', 'S3 %', 'S4 %', 'S5 %',
        'PFA I %', 'PFA II %', 'PFA III %', 'PFA IV %', 'PFA V %'
    ]
    for col in OPERATIVE_COLS + dettaglio_cols:
        df_base[col] = 0.0 if '€' in col or '%' in col else 0
        
    df_base['[N] Sc. Fattura %'] = 0.0
    df_base['[N] Contratto %'] = 0.0

    if 'rinnovi_df' not in st.session_state:
        st.session_state.rinnovi_df = df_base.copy()

    tab_simulazione, tab_risultati, tab_esplosione, tab_storico_rinnovi = st.tabs([
        "1. Master Grid (Input Dati)", 
        "2. Analisi Ponderata & Spazio Promo",
        "3. Esplosione Sconti (Dettaglio)",
        "4. Proposte Rinnovi Salvate 📁"
    ])

    with tab_simulazione:
        st.markdown("#### Griglia di Simulazione Contrattuale")
        filtro_vista = st.radio("Filtra Referenze in Tabella:", ["Tutte le Referenze", "Solo con Volumi > 0", "Sotto Soglia (Allarme Rosso)"], horizontal=True, key="filtro_vista_rinnovi")
        
        with st.container(border=True):
            st.markdown("**Condizioni Logistiche e Finanziarie (Applicate a tutte le referenze)**")
            col_glob1, col_glob2 = st.columns(2)
            with col_glob1:
                st.number_input("Sconto Carico Logistica (%)", min_value=0.0, max_value=100.0, step=0.5, key="global_carico")
            with col_glob2:
                st.number_input("Sconto Pagamento (%)", min_value=0.0, max_value=100.0, step=0.5, key="global_pagamento")
        
        df_kpi = st.session_state.rinnovi_df[st.session_state.rinnovi_df['[N+1] Volumi'] > 0].copy()
        tot_delta_perc = 0.0
        if not df_kpi.empty:
            molt_glob = (1 - st.session_state.global_carico/100) * (1 - st.session_state.global_pagamento/100)
            nn_n = df_kpi['[N] Listino €'] * (1 - df_kpi['[N] Sc. Fattura %']/100) * molt_glob * (1 - df_kpi['[N] Contratto %']/100)
            nn_n1 = df_kpi['[N+1] Listino €'] * (1 - df_kpi['[N+1] Sc. Fattura %']/100) * molt_glob * (1 - df_kpi['[N+1] Contratto %']/100)
            fatt_n = (nn_n * df_kpi['[N+1] Volumi']).sum()
            fatt_n1 = (nn_n1 * df_kpi['[N+1] Volumi']).sum()
            if fatt_n > 0:
                tot_delta_perc = ((fatt_n1 - fatt_n) / fatt_n) * 100

        col_up3, col_up4 = st.columns(2)
        with col_up3:
            with st.container(border=True):
                st.markdown("**Esporta Template Simulazione**")
                buf_sim = io.BytesIO()
                st.session_state.rinnovi_df[['ean', 'Categoria', 'Sub-Categoria', 'Prodotto', 'Minimo Net Net €'] + OPERATIVE_COLS].to_excel(buf_sim, index=False)
                st.download_button("Scarica Tabella Simulazione", buf_sim.getvalue(), "Template_Simulazione.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            st.metric("Variazione Totale Ponderata (%)", fmt_it(tot_delta_perc, 2, is_pct=True, sign=True))

        with col_up4:
            with st.container(border=True):
                st.markdown("**Importa Dati Simulazione**")
                up_sim = st.file_uploader("Carica Excel Simulazione", type=['xlsx'], key="up_sim")
                if up_sim:
                    df_up_sim = pd.read_excel(up_sim, dtype={'ean': str})
                    df_up_sim['ean'] = df_up_sim['ean'].astype(str).str.zfill(13)
                    df_temp = st.session_state.rinnovi_df.copy()
                    for col in OPERATIVE_COLS + ['Minimo Net Net €']:
                        if col in df_up_sim.columns:
                            mapping = df_up_sim.set_index('ean')[col].to_dict()
                            df_temp[col] = df_temp['ean'].map(mapping).fillna(df_temp[col])
                    st.session_state.rinnovi_df = df_temp
                    st.rerun()

        df_display = st.session_state.rinnovi_df.copy()
        moltiplicatore_globale = (1 - st.session_state.global_carico/100) * (1 - st.session_state.global_pagamento/100)
        
        df_display['Net Net [N] €'] = df_display['[N] Listino €'] * (1 - df_display['[N] Sc. Fattura %']/100) * moltiplicatore_globale * (1 - df_display['[N] Contratto %']/100)
        df_display['Net Net [N+1] €'] = df_display['[N+1] Listino €'] * (1 - df_display['[N+1] Sc. Fattura %']/100) * moltiplicatore_globale * (1 - df_display['[N+1] Contratto %']/100)
        df_display['Delta Assoluto €'] = df_display['Net Net [N+1] €'] - df_display['Net Net [N] €']
        
        def calc_max_promo(row):
            nn_base = row['Net Net [N+1] €']
            floor = row['Minimo Net Net €']
            if nn_base > 0 and nn_base > floor:
                return (1 - (floor / nn_base)) * 100
            return 0.0
            
        df_display['Sc. Promo MAX [N+1] %'] = df_display.apply(calc_max_promo, axis=1)
        
        if filtro_vista == "Solo con Volumi > 0":
            df_display = df_display[df_display['[N+1] Volumi'] > 0]
        elif filtro_vista == "Sotto Soglia (Allarme Rosso)":
            df_display = df_display[df_display['Net Net [N+1] €'] < df_display['Minimo Net Net €']]

        col_config = {
            "Prodotto": st.column_config.TextColumn("Prodotto", disabled=True),
            "[N] Listino €": st.column_config.NumberColumn("[N] Listino €", format="€ %.2f", disabled=True),
            "[N+1] Volumi": st.column_config.NumberColumn("[N+1] Volumi", step=100),
            "[N+1] Listino €": st.column_config.NumberColumn("[N+1] Listino €", format="€ %.2f", step=0.1),
            "[N+1] Sc. Fattura %": st.column_config.NumberColumn("[N+1] Sc. Fattura %", format="%.2f %%", step=0.5),
            "[N+1] Contratto %": st.column_config.NumberColumn("[N+1] Contratto %", format="%.2f %%", step=0.5),
            "Net Net [N] €": st.column_config.NumberColumn("Net Net [N] €", format="€ %.3f", disabled=True),
            "Net Net [N+1] €": st.column_config.NumberColumn("Net Net [N+1] €", format="€ %.3f", disabled=True),
            "Minimo Net Net €": st.column_config.NumberColumn("Minimo Net Net €", format="€ %.2f", step=0.05),
            "Sc. Promo MAX [N+1] %": st.column_config.NumberColumn("Sc. Promo MAX [N+1] %", format="%.2f %%", disabled=True),
            "Delta Assoluto €": st.column_config.NumberColumn("Delta Assoluto €", format="€ %+.3f", disabled=True),
        }
        
        cols_to_edit = ['Prodotto', '[N] Listino €', '[N+1] Volumi', '[N+1] Listino €', '[N+1] Sc. Fattura %', '[N+1] Contratto %', 'Net Net [N] €', 'Net Net [N+1] €', 'Minimo Net Net €', 'Sc. Promo MAX [N+1] %', 'Delta Assoluto €']
        
        with st.form("form_simulazione"):
            st.markdown("##### ⚡ Azioni Rapide e Aggiornamento Massivo")
            
            col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns([2, 2, 1.5, 2, 2])
            subcat_uniche = sorted(st.session_state.rinnovi_df['Sub-Categoria'].unique().tolist())
            
            with col_m1:
                cat_mass = st.selectbox("1. Scegli Categoria", ["Tutto l'Assortimento"] + subcat_uniche)
            with col_m2:
                col_mass = st.selectbox("2. Scegli Parametro", [
                    "Aumento Listino Base (%)",
                    "[N+1] Sc. Fattura %", 
                    "[N+1] Contratto %"
                ])
            with col_m3:
                val_mass = st.number_input("3. Valore (%)", min_value=-100.0, max_value=100.0, step=0.5, format="%.2f")
            with col_m4:
                st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
                btn_mass = st.form_submit_button("⚡ Applica Valore", type="secondary")
            with col_m5:
                st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
                submit_sim = st.form_submit_button("🔄 Calcola Simulazione", type="primary")
                
            st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
            
            df_sim_edited = st.data_editor(
                df_display[cols_to_edit],
                column_config=col_config,
                hide_index=True,
                use_container_width=False, 
                height=600,
                key="editor_simulazione_nativa"
            )
            
        if submit_sim or btn_mass:
            for i, idx in enumerate(df_display.index):
                for col in ['[N+1] Volumi', '[N+1] Listino €', '[N+1] Sc. Fattura %', '[N+1] Contratto %', 'Minimo Net Net €']:
                    st.session_state.rinnovi_df.at[idx, col] = df_sim_edited.iloc[i][col]
            
            if btn_mass:
                for idx, row in st.session_state.rinnovi_df.iterrows():
                    if cat_mass == "Tutto l'Assortimento" or row['Sub-Categoria'] == cat_mass:
                        if col_mass == "Aumento Listino Base (%)":
                            listino_n = row['[N] Listino €']
                            if listino_n > 0:
                                nuovo_listino = listino_n * (1 + (val_mass / 100))
                                nuovo_listino_arr = float(Decimal(str(nuovo_listino)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP))
                                st.session_state.rinnovi_df.at[idx, '[N+1] Listino €'] = nuovo_listino_arr
                        else:
                            st.session_state.rinnovi_df.at[idx, col_mass] = val_mass
                        
            st.rerun()

    with tab_risultati:
        df_active = st.session_state.rinnovi_df[st.session_state.rinnovi_df['[N+1] Volumi'] > 0].copy()
        
        if df_active.empty:
            st.warning("Nessuna referenza attiva. Inserisci dei volumi nella Master Grid.")
        else:
            moltiplicatore_globale = (1 - st.session_state.global_carico/100) * (1 - st.session_state.global_pagamento/100)
            
            df_active['Net Net [N] €'] = df_active['[N] Listino €'] * (1 - df_active['[N] Sc. Fattura %']/100) * moltiplicatore_globale * (1 - df_active['[N] Contratto %']/100)
            df_active['Net Net [N+1] €'] = df_active['[N+1] Listino €'] * (1 - df_active['[N+1] Sc. Fattura %']/100) * moltiplicatore_globale * (1 - df_active['[N+1] Contratto %']/100)
            
            df_active['Fatturato_N'] = df_active['Net Net [N] €'] * df_active['[N+1] Volumi'] 
            df_active['Fatturato_N1'] = df_active['Net Net [N+1] €'] * df_active['[N+1] Volumi']
            df_active['Valore_Floor_Totale_N1'] = df_active['Minimo Net Net €'] * df_active['[N+1] Volumi']
            
            df_cat = df_active.groupby('Categoria').agg(
                Volumi_N1=('[N+1] Volumi', 'sum'),
                Fatturato_N=('Fatturato_N', 'sum'),
                Fatturato_N1=('Fatturato_N1', 'sum'),
                Valore_Floor_Totale_N1=('Valore_Floor_Totale_N1', 'sum')
            ).reset_index()
            
            df_cat['Net Net Pond. [N] €'] = df_cat.apply(lambda x: x['Fatturato_N'] / x['Volumi_N1'] if x['Volumi_N1'] > 0 else 0, axis=1)
            df_cat['Net Net Pond. [N+1] €'] = df_cat.apply(lambda x: x['Fatturato_N1'] / x['Volumi_N1'] if x['Volumi_N1'] > 0 else 0, axis=1)
            df_cat['Floor Pond. €'] = df_cat.apply(lambda x: x['Valore_Floor_Totale_N1'] / x['Volumi_N1'] if x['Volumi_N1'] > 0 else 0, axis=1)
            
            df_cat['Delta %'] = df_cat.apply(lambda x: ((x['Net Net Pond. [N+1] €'] - x['Net Net Pond. [N] €']) / x['Net Net Pond. [N] €'] * 100) if x['Net Net Pond. [N] €'] > 0 else 0, axis=1)
            df_cat['Allarme'] = df_cat['Net Net Pond. [N+1] €'] < df_cat['Floor Pond. €']
            
            tot_vol_n1 = df_active['[N+1] Volumi'].sum()
            tot_net_n = df_active['Fatturato_N'].sum() / tot_vol_n1 if tot_vol_n1 > 0 else 0
            tot_net_n1 = df_active['Fatturato_N1'].sum() / tot_vol_n1 if tot_vol_n1 > 0 else 0
            tot_floor_n1 = df_active['Valore_Floor_Totale_N1'].sum() / tot_vol_n1 if tot_vol_n1 > 0 else 0
            tot_delta_perc = ((tot_net_n1 - tot_net_n) / tot_net_n * 100) if tot_net_n > 0 else 0
            
            st.markdown("#### KPI Totali Cliente (Media Ponderata)")
            col_k1, col_k2, col_k3, col_k4 = st.columns(4)
            col_k1.metric("Volumi Totali [N+1]", fmt_it(tot_vol_n1, 0))
            col_k2.metric("Net-Net Pond. [N]", fmt_it(tot_net_n, 3, is_euro=True))
            col_k3.metric("Net-Net Pond. [N+1]", fmt_it(tot_net_n1, 3, is_euro=True), fmt_it(tot_net_n1 - tot_net_n, 3, is_euro=True, sign=True))
            col_k4.metric("Variazione Totale (%)", fmt_it(tot_delta_perc, 2, is_pct=True, sign=True))
            
            st.divider()
            
            st.markdown("#### 1. Sub-Totali per Categoria (Media Ponderata)")
            
            def highlight_cat(row):
                if row['Allarme']: return ['background-color: #FEF2F2; color: #991B1B; font-weight: bold'] * len(row)
                if row['Delta %'] < 0: return ['color: #D97706'] * len(row)
                return [''] * len(row)

            df_cat_disp = df_cat[['Categoria', 'Volumi_N1', 'Net Net Pond. [N] €', 'Net Net Pond. [N+1] €', 'Delta %', 'Floor Pond. €', 'Allarme']]
            
            st.dataframe(
                df_cat_disp.style.apply(highlight_cat, axis=1).format({
                    'Volumi_N1': lambda x: fmt_it(x, 0),
                    'Net Net Pond. [N] €': lambda x: fmt_it(x, 3, is_euro=True), 
                    'Net Net Pond. [N+1] €': lambda x: fmt_it(x, 3, is_euro=True), 
                    'Delta %': lambda x: fmt_it(x, 2, is_pct=True, sign=True), 
                    'Floor Pond. €': lambda x: fmt_it(x, 3, is_euro=True)
                }),
                column_config={"Allarme": "Sotto Floor!"},
                hide_index=True, use_container_width=True
            )
            st.download_button("📥 Scarica Tabella Categorie (Excel)", to_excel_bytes(df_cat_disp), "Analisi_Categorie.xlsx", key="down_cat_rinnovi")
            
            st.divider()
            
            st.markdown("#### 2. Dettaglio Referenze (SKU)")
            
            df_active['Delta Listino €'] = df_active['[N+1] Listino €'] - df_active['[N] Listino €']
            df_active['Delta Listino %'] = df_active.apply(lambda x: ((x['[N+1] Listino €'] - x['[N] Listino €']) / x['[N] Listino €'] * 100) if x['[N] Listino €'] > 0 else 0, axis=1)
            df_active['Delta Assoluto €'] = df_active['Net Net [N+1] €'] - df_active['Net Net [N] €']
            df_active['Delta %'] = df_active.apply(lambda x: ((x['Net Net [N+1] €'] - x['Net Net [N] €']) / x['Net Net [N] €'] * 100) if x['Net Net [N] €'] > 0 else 0, axis=1)
            
            df_active['Spazio Promo %'] = df_active.apply(lambda x: ((1 - (x['Minimo Net Net €'] / x['Net Net [N+1] €'])) * 100) if x['Net Net [N+1] €'] > x['Minimo Net Net €'] else 0, axis=1)
            df_active['Spazio Promo €'] = df_active['Net Net [N+1] €'] - df_active['Minimo Net Net €']
            df_active['Allarme'] = df_active['Net Net [N+1] €'] < df_active['Minimo Net Net €']
            
            cols_sku_disp = [
                'Sub-Categoria', 'Prodotto', 
                '[N] Listino €', '[N+1] Listino €', 'Delta Listino €', 'Delta Listino %',
                'Net Net [N] €', 'Net Net [N+1] €', 'Delta Assoluto €', 'Delta %', 
                'Minimo Net Net €', 'Spazio Promo €', 'Spazio Promo %', 'Allarme'
            ]
            
            st.dataframe(
                df_active[cols_sku_disp].style.apply(highlight_cat, axis=1).format({
                    '[N] Listino €': lambda x: fmt_it(x, 2, is_euro=True),
                    '[N+1] Listino €': lambda x: fmt_it(x, 2, is_euro=True),
                    'Delta Listino €': lambda x: fmt_it(x, 2, is_euro=True, sign=True),
                    'Delta Listino %': lambda x: fmt_it(x, 2, is_pct=True, sign=True),
                    'Net Net [N] €': lambda x: fmt_it(x, 3, is_euro=True), 
                    'Net Net [N+1] €': lambda x: fmt_it(x, 3, is_euro=True), 
                    'Delta Assoluto €': lambda x: fmt_it(x, 3, is_euro=True, sign=True),
                    'Delta %': lambda x: fmt_it(x, 2, is_pct=True, sign=True), 
                    'Minimo Net Net €': lambda x: fmt_it(x, 3, is_euro=True),
                    'Spazio Promo €': lambda x: fmt_it(x, 3, is_euro=True, sign=True),
                    'Spazio Promo %': lambda x: fmt_it(x, 2, is_pct=True)
                }),
                column_config={
                    "Allarme": "Sotto Floor!",
                    "Delta Assoluto €": "Delta Net Net €",
                    "Delta %": "Delta Net Net %"
                },
                hide_index=True, use_container_width=True
            )
            st.download_button("📥 Scarica Dettaglio Referenze (Excel)", to_excel_bytes(df_active[cols_sku_disp]), "Dettaglio_Referenze.xlsx", key="down_sku_rinnovi")

    with tab_esplosione:
        st.markdown("#### Esplosione Sconti (Dettaglio)")
        st.markdown("Una volta definiti i target aggregati nel Tab 1, usa questa griglia per spacchettare gli sconti nelle singole voci contrattuali. Modifica i dati e premi **Calcola e Verifica Sconti**.")
        
        df_explode = st.session_state.rinnovi_df[st.session_state.rinnovi_df['[N+1] Volumi'] > 0].copy()
        
        if df_explode.empty:
            st.warning("Nessuna referenza attiva.")
        else:
            def check_fattura(row):
                p = 1.0
                for s in ['S1 %', 'S2 %', 'S3 %', 'S4 %', 'S5 %']:
                    p *= (1 - (row[s] / 100))
                return (1 - p) * 100
                
            def check_pfa(row):
                return row['PFA I %'] + row['PFA II %'] + row['PFA III %'] + row['PFA IV %'] + row['PFA V %']

            df_explode['Check Sc. Fattura %'] = df_explode.apply(check_fattura, axis=1)
            df_explode['Check PFA %'] = df_explode.apply(check_pfa, axis=1)
            
            df_explode['Delta Fattura'] = df_explode['[N+1] Sc. Fattura %'] - df_explode['Check Sc. Fattura %']
            df_explode['Delta PFA'] = df_explode['[N+1] Contratto %'] - df_explode['Check PFA %']

            col_config_exp = {
                "Prodotto": st.column_config.TextColumn("Prodotto", disabled=True),
                "[N+1] Sc. Fattura %": st.column_config.NumberColumn("Target Fattura %", format="%.2f %%", disabled=True),
                "Check Sc. Fattura %": st.column_config.NumberColumn("Somma Geom. %", format="%.2f %%", disabled=True),
                "Delta Fattura": st.column_config.NumberColumn("Diff. Fattura", format="%+.2f", disabled=True),
                "[N+1] Contratto %": st.column_config.NumberColumn("Target PFA %", format="%.2f %%", disabled=True),
                "Check PFA %": st.column_config.NumberColumn("Somma Algeb. %", format="%.2f %%", disabled=True),
                "Delta PFA": st.column_config.NumberColumn("Diff. PFA", format="%+.2f", disabled=True),
            }
            
            for col in dettaglio_cols:
                col_config_exp[col] = st.column_config.NumberColumn(col, format="%.2f %%", step=0.5)

            cols_to_edit_exp = [
                'Prodotto', '[N+1] Sc. Fattura %', 'Check Sc. Fattura %', 'Delta Fattura',
                'S1 %', 'S2 %', 'S3 %', 'S4 %', 'S5 %',
                '[N+1] Contratto %', 'Check PFA %', 'Delta PFA',
                'PFA I %', 'PFA II %', 'PFA III %', 'PFA IV %', 'PFA V %'
            ]
            
            col_spacer, col_btn_exp = st.columns([3, 1])
            with col_btn_exp:
                st.download_button("📥 Scarica Esplosione Sconti (Excel)", to_excel_bytes(df_explode[cols_to_edit_exp]), "Esplosione_Sconti.xlsx", key="down_explode_rinnovi")

            with st.form("form_esplosione"):
                # --- STRUMENTO DI VARIAZIONE MASSIVA ---
                st.markdown("##### ⚡ Azioni Rapide e Aggiornamento Massivo")
                col_m1, col_m2, col_m3, col_m4 = st.columns([2.5, 2.5, 2, 2])
                subcat_uniche = sorted(st.session_state.rinnovi_df['Sub-Categoria'].unique().tolist())
                
                with col_m1:
                    cat_mass_exp = st.selectbox("1. Scegli Categoria", ["Tutto l'Assortimento"] + subcat_uniche, key="cat_mass_exp")
                with col_m2:
                    param_mass_exp = st.selectbox("2. Scegli Sconto / Voce PFA", [
                        "S1 %", "S2 %", "S3 %", "S4 %", "S5 %",
                        "PFA I %", "PFA II %", "PFA III %", "PFA IV %", "PFA V %"
                    ], key="param_mass_exp")
                with col_m3:
                    val_mass_exp = st.number_input("3. Valore (%)", min_value=0.0, max_value=100.0, step=0.5, format="%.2f", key="val_mass_exp")
                with col_m4:
                    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
                    btn_mass_exp = st.form_submit_button("⚡ Applica Massivo")
                    
                st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
                
                col_btn_calc, col_btn_align = st.columns(2)
                with col_btn_calc:
                    submit_exp = st.form_submit_button("🔄 Calcola e Verifica Sconti (Manuale)", type="primary")
                with col_btn_align:
                    align_exp = st.form_submit_button("🪄 Allinea Sconti Automaticamente", type="secondary", help="Calcola la differenza matematica e la inserisce in S5 (Fattura) e PFA V (Fuori Fattura) per centrare il target.")
                
                df_exp_edited = st.data_editor(
                    df_explode[cols_to_edit_exp],
                    column_config=col_config_exp,
                    hide_index=True,
                    use_container_width=False,
                    height=600,
                    key="editor_esplosione"
                )
                
            if submit_exp or align_exp or btn_mass_exp:
                 # Salvataggio preventivo modifiche manuali
                 for i, idx in enumerate(df_explode.index):
                     for col in dettaglio_cols:
                         if col in df_exp_edited.columns:
                             st.session_state.rinnovi_df.at[idx, col] = df_exp_edited.iloc[i][col]
                             
                 # Applicazione del valore massivo se richiesto
                 if btn_mass_exp:
                     df_temp = st.session_state.rinnovi_df.copy()
                     for idx, row in df_temp.iterrows():
                         if row['[N+1] Volumi'] > 0:
                             if cat_mass_exp == "Tutto l'Assortimento" or row['Sub-Categoria'] == cat_mass_exp:
                                 df_temp.at[idx, param_mass_exp] = val_mass_exp
                     st.session_state.rinnovi_df = df_temp
                             
                 # Allineamento automatico se richiesto
                 if align_exp:
                     df_temp = st.session_state.rinnovi_df.copy()
                     for idx, row in df_temp[df_temp['[N+1] Volumi'] > 0].iterrows():
                         target_fatt = row['[N+1] Sc. Fattura %']
                         p_parziale = 1.0
                         for s in ['S1 %', 'S2 %', 'S3 %', 'S4 %']:
                             p_parziale *= (1 - (row[s] / 100))
                         
                         if p_parziale > 0:
                             s5_req = (1 - ((1 - target_fatt/100) / p_parziale)) * 100
                             df_temp.at[idx, 'S5 %'] = round(s5_req, 2)
                         
                         target_pfa = row['[N+1] Contratto %']
                         pfa_parziale = row['PFA I %'] + row['PFA II %'] + row['PFA III %'] + row['PFA IV %']
                         df_temp.at[idx, 'PFA V %'] = round(target_pfa - pfa_parziale, 2)
                     
                     st.session_state.rinnovi_df = df_temp
                     
                 st.rerun()

    with tab_storico_rinnovi:
        st.markdown("#### Gestione Proposte dei Rinnovi Contrattuali Salvate")
        
        df_saved = pd.read_sql_query("""
            SELECT id, data_salvataggio, nome_proposta, gruppo_macro, sottogruppo, associato_insegna, global_carico, global_pagamento 
            FROM proposte_rinnovi 
            ORDER BY data_salvataggio DESC
        """, conn)
        
        if df_saved.empty:
            st.info("Nessuna proposta di rinnovo salvata al momento. Usa il pannello 'Azioni Scenario' in alto per memorizzare la simulazione corrente.")
        else:
            st.dataframe(
                df_saved,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "id": "ID Scenario",
                    "data_salvataggio": st.column_config.DatetimeColumn("Data Salvataggio", format="DD/MM/YYYY HH:mm"),
                    "nome_proposta": "Nome Proposta",
                    "gruppo_macro": "Gruppo Macro",
                    "sottogruppo": "Sottogruppo",
                    "associato_insegna": "Insegna Locale",
                    "global_carico": st.column_config.NumberColumn("Sconto Carico %", format="%.2f %%"),
                    "global_pagamento": st.column_config.NumberColumn("Sconto Pagamento %", format="%.2f %%"),
                }
            )
            
            col_load_left, col_load_right = st.columns(2)
            
            scenari_opzioni = {f"{r['nome_proposta']} ({r['associato_insegna'] or r['gruppo_macro']}) [ID: {r['id']}]": r['id'] for _, r in df_saved.iterrows()}
            
            with col_load_left:
                with st.container(border=True):
                    st.markdown("**🔄 Ripristina Proposta**")
                    sel_scenario_text = st.selectbox("Seleziona scenario da ricaricare nel simulatore", list(scenari_opzioni.keys()), key="select_scenario_load_key")
                    scenario_id = scenari_opzioni[sel_scenario_text]
                    
                    if st.button("CARICA PROPOSTA SELEZIONATA 🔄", type="primary", use_container_width=True, help="Sovrascrive lo scenario corrente ripristinando la simulazione caricata."):
                        cursor.execute("SELECT gruppo_macro, sottogruppo, associato_insegna, global_carico, global_pagamento, dati_json FROM proposte_rinnovi WHERE id=?", (scenario_id,))
                        res_scen = cursor.fetchone()
                        if res_scen:
                            st.session_state.rinnovi_gruppo = res_scen[0]
                            st.session_state.rinnovi_sottogruppo = res_scen[1] or ""
                            st.session_state.rinnovi_insegna = res_scen[2] or ""
                            st.session_state.global_carico = float(res_scen[3])
                            st.session_state.global_pagamento = float(res_scen[4])
                            
                            df_restored = pd.read_json(io.StringIO(res_scen[5]))
                            if 'ean' in df_restored.columns:
                                df_restored['ean'] = df_restored['ean'].astype(str).str.replace(r'\.0$', '', regex=True).str.zfill(13)
                                
                            st.session_state.rinnovi_df = df_restored
                            st.success(f"Scenario ID {scenario_id} ripristinato con successo.")
                            st.rerun()
                            
            with col_load_right:
                with st.container(border=True):
                    st.markdown("**❌ Elimina Proposta**")
                    sel_scenario_del_text = st.selectbox("Seleziona scenario da rimuovere", list(scenari_opzioni.keys()), key="select_scenario_del_key")
                    scenario_del_id = scenari_opzioni[sel_scenario_del_text]
                    
                    if st.button("ELIMINA PROPOSTA SELEZIONATA ❌", type="secondary", use_container_width=True, help="Elimina in modo definitivo lo scenario selezionato dal DB."):
                        cursor.execute("DELETE FROM proposte_rinnovi WHERE id=?", (scenario_del_id,))
                        conn.commit()
                        st.warning(f"Scenario ID {scenario_del_id} rimosso con successo.")
                        st.rerun()

    conn.close()

# ==========================================
# SCHEDA: STORICO PROMOZIONI
# ==========================================
elif menu == "Storico Promozioni":
    st.title("Storico Promozioni (CRM Commerciale)")
    st.markdown("Archivio delle simulazioni salvate. Filtra per cliente o stato per recuperare le trattative passate.")
    
    conn = sqlite3.connect(DB_FILE)
    
    with st.container(border=True):
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            filtro_stato = st.selectbox("Filtra per Stato", ["Tutti", "Confermata", "Proposta"])
        with col_f2:
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT gruppo_macro FROM storico_promo ORDER BY gruppo_macro")
            gruppi_storico = ["Tutti"] + [r[0] for r in cursor.fetchall()]
            filtro_gruppo = st.selectbox("Filtra per Gruppo", gruppi_storico)
        with col_f3:
            if filtro_gruppo != "Tutti":
                cursor.execute("SELECT DISTINCT associato_insegna FROM storico_promo WHERE gruppo_macro=? ORDER BY associato_insegna", (filtro_gruppo,))
                insegne_storico = ["Tutte"] + [r[0] for r in cursor.fetchall()]
            else:
                insegne_storico = ["Tutte"]
            filtro_insegna = st.selectbox("Filtra per Insegna", insegne_storico)

    query = """
        SELECT id, data_salvataggio, stato_promo, gruppo_macro, associato_insegna, ean, descrizione_commerciale, 
               sell_in_dal, sell_in_al, sell_out_dal, sell_out_al,
               listino_r, sconto_y, sconto_z, sconto_aa, min_net_net_g, net_net_am, net_net_post_promo, 
               volumi_stimati, contributo_fisso, contributo_pezzo, costo_totale_extra, note 
        FROM storico_promo WHERE 1=1
    """
    params = []
    
    if filtro_stato != "Tutti":
        query += " AND stato_promo = ?"
        params.append(filtro_stato)
    if filtro_gruppo != "Tutti":
        query += " AND gruppo_macro = ?"
        params.append(filtro_gruppo)
    if filtro_insegna != "Tutte":
        query += " AND associato_insegna = ?"
        params.append(filtro_insegna)
        
    query += " ORDER BY data_salvataggio DESC"
    
    df_storico = pd.read_sql_query(query, conn, params=params)
    
    st.dataframe(
        df_storico.drop(columns=['ean', 'contributo_fisso', 'contributo_pezzo']), 
        use_container_width=True, 
        hide_index=True,
        column_config={
            "id": "ID",
            "data_salvataggio": st.column_config.DatetimeColumn("Data Salvataggio", format="DD/MM/YYYY HH:mm"),
            "stato_promo": "Stato",
            "gruppo_macro": "Gruppo",
            "associato_insegna": "Insegna",
            "descrizione_commerciale": "Prodotto",
            "sell_in_dal": st.column_config.DateColumn("Inizio Sell-In", format="DD/MM/YYYY"),
            "sell_in_al": st.column_config.DateColumn("Fine Sell-In", format="DD/MM/YYYY"),
            "sell_out_dal": st.column_config.DateColumn("Inizio Sell-Out", format="DD/MM/YYYY"),
            "sell_out_al": st.column_config.DateColumn("Fine Sell-Out", format="DD/MM/YYYY"),
            "listino_r": st.column_config.NumberColumn("Listino R", format="€ %.2f"),
            "sconto_z": st.column_config.NumberColumn("Sc. Z (%)", format="%.2f %%"),
            "sconto_aa": st.column_config.NumberColumn("Sc. AA (€)", format="€ %.2f"),
            "min_net_net_g": st.column_config.NumberColumn("Min Net Net (G)", format="€ %.3f"),
            "net_net_am": st.column_config.NumberColumn("Net Net (AM)", format="€ %.3f"),
            "net_net_post_promo": st.column_config.NumberColumn("Net Net Post-Promo", format="€ %.3f"),
            "volumi_stimati": "Volumi (Pz)",
            "costo_totale_extra": st.column_config.NumberColumn("Costo Extra", format="€ %.2f"),
            "note": "Note"
        }
    )
    
    col_export, col_clone, col_delete = st.columns(3)
    
    with col_export:
        st.markdown("#### Esporta Dati")
        if not df_storico.empty:
            buffer_storico = io.BytesIO()
            with pd.ExcelWriter(buffer_storico, engine='openpyxl') as writer:
                df_storico.to_excel(writer, index=False, sheet_name="Storico_Promo")
                
            st.download_button(
                label="SCARICA ESTRAZIONE (Excel)",
                data=buffer_storico.getvalue(),
                file_name=f"Storico_Promozioni_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            st.info("Nessuna promozione salvata.")
            
    with col_clone:
        with st.container(border=True):
            st.markdown("#### Clona Promozione")
            if not df_storico.empty:
                id_to_clone = st.selectbox("Seleziona l'ID da ricaricare:", df_storico['id'].tolist(), key="clone_id")
                if st.button("CLONA NEL SIMULATORE", use_container_width=True):
                    promo_data = df_storico[df_storico['id'] == id_to_clone].iloc[0]
                    st.session_state['widget_gruppo'] = promo_data['gruppo_macro']
                    st.session_state['widget_insegna'] = promo_data['associato_insegna']
                    st.session_state['clone_ean_pending'] = promo_data['ean']
                    st.session_state['widget_metodo'] = "B. Tentativi Spot Manuali (Immissione Sconto Promo libera)"
                    st.session_state['widget_z'] = float(promo_data['sconto_z'])
                    st.session_state['widget_aa'] = float(promo_data['sconto_aa'])
                    st.session_state['widget_volumi'] = int(promo_data['volumi_stimati'])
                    st.session_state['widget_fisso'] = float(promo_data['contributo_fisso'])
                    st.session_state['widget_pezzo'] = float(promo_data['contributo_pezzo'])
                    
                    st.session_state.go_to_menu = "Simulatore Offerte"
                    st.rerun()
            else:
                st.write("Nessun record disponibile.")

    with col_delete:
        with st.container(border=True):
            st.markdown("#### Elimina Record")
            if not df_storico.empty:
                id_to_delete = st.selectbox("Seleziona l'ID da eliminare:", df_storico['id'].tolist(), key="del_id")
                if st.button("ELIMINA DEFINITIVAMENTE", use_container_width=True):
                    try:
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM storico_promo WHERE id=?", (id_to_delete,))
                        conn.commit()
                        st.success(f"ID {id_to_delete} eliminato.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Errore: {e}")
            else:
                st.write("Nessun record disponibile.")
        
    conn.close()

# ==========================================
# SCHEDA 2: ANAGRAFICA GDO (CLIENTI)
# ==========================================
elif menu == "Anagrafica GDO (Clienti)":
    st.title("Anagrafica GDO Italiana")
    st.markdown("Attiva o disattiva le insegne per renderle visibili nel Simulatore Offerte e negli Accordi Locali.")
    
    conn = sqlite3.connect(DB_FILE)
    
    df_gdo = pd.read_sql_query("SELECT id, gruppo_macro, sottogruppo, associato_insegna, attivo FROM struttura_gdo ORDER BY gruppo_macro, sottogruppo, associato_insegna", conn)
    df_gdo['attivo'] = df_gdo['attivo'].astype(bool)
    
    with st.container(border=True):
        st.markdown("#### Gestione Stato Insegne")
        edited_gdo = st.data_editor(
            df_gdo,
            hide_index=True,
            use_container_width=True,
            disabled=["id", "gruppo_macro", "sottogruppo", "associato_insegna"]
        )
        
        if st.button("SALVA STATO INSEGNE", type="primary"):
            cursor = conn.cursor()
            try:
                with conn:
                    for _, r in edited_gdo.iterrows():
                        cursor.execute("UPDATE struttura_gdo SET attivo=? WHERE id=?", (1 if r['attivo'] else 0, r['id']))
                st.success("Stato insegne aggiornato correttamente.")
                st.rerun()
            except Exception as e:
                st.error(f"Errore durante il salvataggio: {e}")
    conn.close()

# ==========================================
# SCHEDA 2: DATI ANAGRAFICI (PRODOTTI E LOGISTICA)
# ==========================================
elif menu == "Dati Anagrafici (Logistica)":
    st.title("Dati Anagrafici - Prodotti e Logistica")
    st.markdown("Gestione dell'anagrafica prodotti (Dati SAP e Logistici). I margini finanziari sono gestiti separatamente.")
    
    conn = sqlite3.connect(DB_FILE)
    
    with st.container(border=True):
        st.markdown("#### Modifica Diretta Anagrafica Master")
        df_prodotti = pd.read_sql_query("""
            SELECT ean, codice_sap, tipo_olio, descrizione_sap, descrizione_commerciale, formato_lt, confezione,
                   pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet, conservazione_mesi, shelf_life_mesi
            FROM anagrafica_master
        """, conn)
        
        edited_prod_df = st.data_editor(
            df_prodotti, 
            num_rows="dynamic", 
            use_container_width=True,
            hide_index=True,
            key="prod_data_editor"
        )
        
        if st.button("SALVA MODIFICHE ANAGRAFICA", type="primary"):
            cursor = conn.cursor()
            try:
                with conn:
                    cursor.execute("DELETE FROM anagrafica_master")
                    for _, r in edited_prod_df.iterrows():
                        def check_nan_float(val):
                            return float(val) if (pd.notna(val) and str(val).strip() != "") else 0.0
                        def check_nan_int(val):
                            return int(float(val)) if (pd.notna(val) and str(val).strip() != "") else 0
                        
                        cursor.execute("""
                        INSERT INTO anagrafica_master (
                            ean, codice_sap, tipo_olio, descrizione_sap, descrizione_commerciale, formato_lt, confezione,
                            pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet, conservazione_mesi, shelf_life_mesi
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            str(r.get("ean")).strip(),
                            str(r.get("codice_sap")).strip(),
                            str(r.get("tipo_olio")).strip(),
                            str(r.get("descrizione_sap")).strip(),
                            str(r.get("descrizione_commerciale")).strip(),
                            check_nan_float(r.get("formato_lt")),
                            str(r.get("confezione")).strip(),
                            check_nan_int(r.get("pezzi_cartone")),
                            check_nan_int(r.get("cartoni_strato")),
                            check_nan_int(r.get("stari_pallet")),
                            check_nan_int(r.get("cartoni_pallet")),
                            check_nan_int(r.get("conservazione_mesi")),
                            check_nan_int(r.get("shelf_life_mesi"))
                        ))
                st.success("Anagrafica aggiornata correttamente.")
                st.rerun()
            except Exception as e:
                st.error(f"Errore durante il salvataggio: {e}")

    col_p1, col_p2 = st.columns(2)
    
    with col_p1:
        with st.container(border=True):
            st.markdown("#### Esportazione Anagrafica")
            st.markdown("Scarica l'anagrafica logistica attuale per lavorarla in Excel.")
            
            buffer_prod_export = io.BytesIO()
            with pd.ExcelWriter(buffer_prod_export, engine='openpyxl') as writer:
                df_prodotti.to_excel(writer, index=False, sheet_name="Anagrafica_SAP")
                
            st.download_button(
                label="Scarica Anagrafica Prodotti (Excel)",
                data=buffer_prod_export.getvalue(),
                file_name=f"Anagrafica_Prodotti_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with col_p2:
        with st.container(border=True):
            st.markdown("#### Importazione Massiva SAP")
            st.markdown("Carica il file Excel per aggiornare l'anagrafica logistica e i guardrail finanziari.")
            uploaded_prod_file = st.file_uploader("Trascina il file Excel Anagrafica (.xlsx)", type=["xlsx"])
            
            if uploaded_prod_file is not None:
                if st.button("Conferma Scrittura Anagrafica"):
                    try:
                        df_prod_import = pd.read_excel(uploaded_prod_file)
                        df_prod_import = DataSanitizer.sanitize_excel_import(df_prod_import)
                        
                        col_map = {
                            "EAN": "ean", "Codice Articolo": "codice_sap", "TIPO OLIO contenuto": "tipo_olio",
                            "Descrizione articolo in SAP": "descrizione_sap", "Descrizione Articolo": "descrizione_commerciale",
                            "Formato (lt)": "formato_lt", "Tipologia Confezione": "confezione",
                            "Pezzi x\nCartone": "pezzi_cartone", "Pezzi x Cartone": "pezzi_cartone",
                            "Cartoni x\nStrato": "cartoni_strato", "Cartoni x Strato": "cartoni_strato",
                            "Strati x Pallet": "strati_pallet", "Cartoni X Pallet": "cartoni_pallet",
                            "Conservazione (Mese)": "conservazione_mesi", "SHELF LIFE (mesi)": "shelf_life_mesi",
                            "Margine Minimo G": "min_net_net_g", "MIN_NET_NET_G": "min_net_net_g", "Soglia Sicurezza G (Euro)": "min_net_net_g"
                        }
                        
                        df_prod_import = df_prod_import.rename(columns=col_map)
                        df_prod_import.columns = [str(c).lower().strip() for c in df_prod_import.columns]
                        
                        if "ean" not in df_prod_import.columns:
                            st.error("Colonna 'EAN' mancante nel file Excel.")
                        else:
                            cursor = conn.cursor()
                            righe_inserite = 0
                            
                            with conn:
                                for idx, row in df_prod_import.iterrows():
                                    ean_val = str(row.get("ean", "")).split('.')[0].zfill(13)
                                    if not ean_val or ean_val == "0000000000000" or ean_val == "nan":
                                        continue
                                        
                                    tipo_olio_raw = str(row.get("tipo_olio", "")).upper().strip()
                                    if tipo_olio_raw == "EXTRA": tipo_olio_raw = "EXTRAVERGINE"

                                    def get_float(col_name, default=0.0):
                                        val = row.get(col_name)
                                        if pd.isna(val) or str(val).strip() == "": return default
                                        try: return float(str(val).replace(',', '.'))
                                        except: return default
                                        
                                    def get_int(col_name, default=0):
                                        val = row.get(col_name)
                                        if pd.isna(val) or str(val).strip() == "": return default
                                        try: return int(float(val))
                                        except: return default

                                    min_g = row.get("min_net_net_g")
                                    if pd.isna(min_g) or str(min_g).strip() == "":
                                        cursor.execute("SELECT min_net_net_g FROM guardrail_aziendali WHERE ean=?", (ean_val,))
                                        res_min = cursor.fetchone()
                                        min_g = res_min[0] if res_min else 0.0
                                    else:
                                        min_g = float(str(min_g).replace(',', '.'))

                                    cursor.execute("""
                                    INSERT OR REPLACE INTO anagrafica_master (
                                        ean, codice_sap, tipo_olio, descrizione_sap, descrizione_commerciale, formato_lt, confezione,
                                        pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet, conservazione_mesi, shelf_life_mesi
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    """, (
                                        ean_val, str(row.get("codice_sap", "")).split('.')[0], tipo_olio_raw,
                                        str(row.get("descrizione_sap", "")), str(row.get("descrizione_commerciale", "")),
                                        get_float("formato_lt"), str(row.get("confezione", "")),
                                        get_int("pezzi_cartone"), get_int("cartoni_strato"), get_int("strati_pallet"),
                                        get_int("cartoni_pallet"), get_int("conservazione_mesi"), get_int("shelf_life_mesi")
                                    ))

                                    cursor.execute("INSERT OR REPLACE INTO guardrail_aziendali (ean, min_net_net_g) VALUES (?, ?)", (ean_val, min_g))
                                    righe_inserite += 1
                            st.success(f"Elaborati {righe_inserite} prodotti.")
                            st.rerun()
                    except Exception as e:
                        st.error(f"Errore durante l'elaborazione: {e}")

    conn.close()

# ==========================================
# SCHEDA 3: BACK-OFFICE CONTRATTI (NAZIONALI)
# ==========================================
elif menu == "Back-Office (Contratti Nazionali)":
    st.title("Back-Office - Contratti Nazionali")
    st.markdown("Gestione degli accordi strutturali (Listini, S1-S5, PFA) a livello di Gruppo e Sottogruppo. **Gli sconti locali si inseriscono nella scheda apposita.**")
    conn = sqlite3.connect(DB_FILE)
    
    # --- INSERIMENTO RAPIDO REGOLA NAZIONALE ---
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT gruppo_macro FROM struttura_gdo WHERE attivo=1 ORDER BY gruppo_macro")
    gruppi_attivi = [r[0] for r in cursor.fetchall()]
    
    cursor.execute("SELECT DISTINCT tipo_olio FROM anagrafica_master")
    categorie_disponibili = [r[0] for r in cursor.fetchall()]
    
    cursor.execute("SELECT ean, descrizione_commerciale FROM anagrafica_master")
    prod_list = [f"{r[1]} [{r[0]}]" for r in cursor.fetchall()]
    
    with st.container(border=True):
        st.markdown("#### ➕ Inserimento Rapido Regola Nazionale")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            ins_gruppo = st.selectbox("Gruppo", gruppi_attivi, key="ins_g")
            
            cursor.execute("""
                SELECT DISTINCT sottogruppo FROM accordi_commerciali 
                WHERE gruppo_macro=? AND sottogruppo != '' AND sottogruppo IS NOT NULL
                UNION
                SELECT DISTINCT sottogruppo FROM struttura_gdo 
                WHERE gruppo_macro=? AND sottogruppo != '' AND sottogruppo IS NOT NULL
                ORDER BY sottogruppo
            """, (ins_gruppo, ins_gruppo))
            sottogruppi_attivi_del_gruppo = [r[0] for r in cursor.fetchall()]
            
            ins_sottogruppo = st.selectbox(
                "Sottogruppo", 
                [""] + sottogruppi_attivi_del_gruppo, 
                key="ins_sg"
            )
            
        with col2:
            ins_livello = st.selectbox("Livello Regola", ["GRUPPO", "SOTTOGRUPPO", "CATEGORIA", "REFERENZA"], key="ins_l")
            
            if ins_livello == "CATEGORIA":
                ins_chiave = st.selectbox("Seleziona Categoria", categorie_disponibili, key="ins_c_cat")
            elif ins_livello == "REFERENZA":
                ins_chiave = st.selectbox("Seleziona Referenza", prod_list, key="ins_c_ref")
            else:
                st.info(f"Regola applicata a tutto il {ins_livello}")
                ins_chiave = ""
            
        with col3:
            ins_parametro = st.selectbox("Parametro da Impostare", [
                "Listino Base (R)", "Sconto 1", "Sconto 2", "Sconto 3", "Sconto 4", "Sconto 5",
                "Oneri Logistica", "Oneri Pagamento", 
                "PFA Voce I", "PFA Voce II", "PFA Voce III", "PFA Voce IV", "PFA Voce V"
            ], key="ins_p")
            ins_valore = st.number_input("Valore", step=0.5, format="%.2f", key="ins_v")
            
        with col4:
            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
            if st.button("Inserisci / Aggiorna Regola", type="primary", use_container_width=True):
                chiave_pulita = ""
                if ins_livello == "CATEGORIA":
                    chiave_pulita = ins_chiave
                elif ins_livello == "REFERENZA":
                    chiave_pulita = ins_chiave.split("[")[-1].replace("]", "").strip() if "[" in ins_chiave else ins_chiave
                
                col_map = {
                    "Listino Base (R)": "listino_r", "Sconto 1": "sconto_1", "Sconto 2": "sconto_2",
                    "Sconto 3": "sconto_3", "Sconto 4": "sconto_4", "Sconto 5": "sconto_5",
                    "Oneri Logistica": "sconto_carico", "Oneri Pagamento": "sconto_pagamento",
                    "PFA Voce I": "voce_contratto_1", "PFA Voce II": "voce_contratto_2", 
                    "PFA Voce III": "voce_contratto_3", "PFA Voce IV": "voce_contratto_4", "PFA Voce V": "voce_contratto_5"
                }
                col_name = col_map[ins_parametro]
                
                try:
                    with conn:
                        cursor.execute("SELECT id FROM accordi_commerciali WHERE gruppo_macro=? AND sottogruppo=? AND associato_insegna='' AND livello=? AND chiave_livello=?", (ins_gruppo, ins_sottogruppo, ins_livello, chiave_pulita))
                        row = cursor.fetchone()
                        if row:
                            cursor.execute(f"UPDATE accordi_commerciali SET {col_name}=? WHERE id=?", (ins_valore, row[0]))
                        else:
                            cursor.execute(f"INSERT INTO accordi_commerciali (gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello, {col_name}) VALUES (?, ?, '', ?, ?, ?)", (ins_gruppo, ins_sottogruppo, ins_livello, chiave_pulita, ins_valore))
                    st.success("Regola salvata!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Errore: {e}")
    # -------------------------------------------
    
    tab_contratti, tab_guardrail = st.tabs(["Gestione Contratti Nazionali", "Gestione Minimi Net Net"])
    
    with tab_contratti:
        with st.container(border=True):
            st.markdown("#### Modifica Diretta dei Contratti in Database")
            
            df_database_editor = pd.read_sql_query("""
                SELECT a.id, a.gruppo_macro, a.sottogruppo, a.livello, a.chiave_livello,
                       CASE 
                            WHEN a.livello = 'REFERENZA' THEN p.descrizione_commerciale 
                            WHEN a.livello = 'CATEGORIA' THEN 'Categoria: ' || a.chiave_livello
                            ELSE 'Contratto Quadro'
                       END as descrizione_prodotto,
                       a.listino_r,
                       a.sconto_1, a.sconto_2, a.sconto_3, a.sconto_4, a.sconto_5,
                       a.sconto_carico, a.sconto_pagamento, a.voce_contratto_1, a.voce_contratto_2, a.voce_contratto_3,
                       a.voce_contratto_4, a.voce_contratto_5
                FROM accordi_commerciali a
                LEFT JOIN anagrafica_master p ON a.chiave_livello = p.ean AND a.livello = 'REFERENZA'
                WHERE a.associato_insegna = '' OR a.associato_insegna IS NULL
                ORDER BY a.gruppo_macro ASC, a.sottogruppo ASC, 
                    CASE a.livello
                        WHEN 'GRUPPO' THEN 1
                        WHEN 'SOTTOGRUPPO' THEN 2
                        WHEN 'CATEGORIA' THEN 3
                        WHEN 'REFERENZA' THEN 4
                        ELSE 5
                    END ASC
            """, conn)
            
            edited_df = st.data_editor(
                df_database_editor, 
                num_rows="dynamic", 
                use_container_width=True,
                hide_index=True,
                disabled=["descrizione_prodotto"],
                key="db_data_editor"
            )
            
            if st.button("SALVA MODIFICHE CONTRATTI NAZIONALI", type="primary"):
                cursor = conn.cursor()
                try:
                    with conn:
                        cursor.execute("DELETE FROM accordi_commerciali WHERE associato_insegna = '' OR associato_insegna IS NULL")
                        for _, r in edited_df.iterrows():
                            def check_nan(val):
                                return float(val) if (pd.notna(val) and str(val).strip() != "") else None
                            
                            cursor.execute("""
                            INSERT OR REPLACE INTO accordi_commerciali (
                                id, gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello, listino_r,
                                sconto_1, sconto_2, sconto_3, sconto_4, sconto_5,
                                sconto_carico, sconto_pagamento, voce_contratto_1, voce_contratto_2, voce_contratto_3,
                                voce_contratto_4, voce_contratto_5
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (
                                check_nan(r.get("id")),
                                str(r.get("gruppo_macro")).upper().strip() if pd.notna(r.get("gruppo_macro")) else "",
                                str(r.get("sottogruppo")).upper().strip() if pd.notna(r.get("sottogruppo")) else "",
                                "", 
                                str(r.get("livello")).upper().strip() if pd.notna(r.get("livello")) else "GRUPPO",
                                str(r.get("chiave_livello")).strip() if pd.notna(r.get("chiave_livello")) else "",
                                check_nan(r.get("listino_r")),
                                check_nan(r.get("sconto_1")), check_nan(r.get("sconto_2")), check_nan(r.get("sconto_3")),
                                check_nan(r.get("sconto_4")), check_nan(r.get("sconto_5")),
                                check_nan(r.get("sconto_carico")), check_nan(r.get("sconto_pagamento")),
                                check_nan(r.get("voce_contratto_1")), check_nan(r.get("voce_contratto_2")), check_nan(r.get("voce_contratto_3")),
                                check_nan(r.get("voce_contratto_4")), check_nan(r.get("voce_contratto_5"))
                            ))
                    st.success("Contratti Nazionali aggiornati correttamente.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Errore durante l'elaborazione: {e}")

        col_b1, col_b2 = st.columns(2)
        
        with col_b1:
            with st.container(border=True):
                st.markdown("#### Esportazione Contratti Nazionali")
                
                query_accordi = """
                SELECT a.gruppo_macro as GRUPPO_MACRO, a.sottogruppo as SOTTOGRUPPO,
                       a.livello as LIVELLO, a.chiave_livello as CHIAVE_LIVELLO,
                       CASE 
                            WHEN a.livello = 'REFERENZA' THEN p.descrizione_commerciale 
                            WHEN a.livello = 'CATEGORIA' THEN 'Accordo di Categoria: ' || a.chiave_livello
                            ELSE 'Contratto Quadro'
                       END as DESCRIZIONE_PRODOTTO,
                       a.listino_r as LISTINO_BASE_R,
                       a.sconto_1 as SCONTO_1, a.sconto_2 as SCONTO_2, a.sconto_3 as SCONTO_3, a.sconto_4 as SCONTO_4, a.sconto_5 as SCONTO_5,
                       a.sconto_carico as SCONTO_CARICO_LOGISTICA, a.sconto_pagamento as SCONTO_PAGAMENTO_AC,
                       a.voce_contratto_1 as PFA_VOCE_I, a.voce_contratto_2 as PFA_VOCE_II,
                       a.voce_contratto_3 as PFA_VOCE_III, a.voce_contratto_4 as PFA_VOCE_IV, a.voce_contratto_5 as PFA_VOCE_V
                FROM accordi_commerciali a
                LEFT JOIN anagrafica_master p ON a.chiave_livello = p.ean AND a.livello = 'REFERENZA'
                WHERE a.associato_insegna = '' OR a.associato_insegna IS NULL
                """
                df_accordi = pd.read_sql_query(query_accordi, conn)
                
                colonne_ordinate = [
                    "GRUPPO_MACRO", "SOTTOGRUPPO", "LIVELLO", "CHIAVE_LIVELLO", "DESCRIZIONE_PRODOTTO",
                    "LISTINO_BASE_R", "SCONTO_1", "SCONTO_2", "SCONTO_3", "SCONTO_4", "SCONTO_5",
                    "SCONTO_CARICO_LOGISTICA", "SCONTO_PAGAMENTO_AC",
                    "PFA_VOCE_I", "PFA_VOCE_II", "PFA_VOCE_III", "PFA_VOCE_IV", "PFA_VOCE_V"
                ]
                df_accordi = df_accordi[colonne_ordinate]
                
                buffer_export = io.BytesIO()
                with pd.ExcelWriter(buffer_export, engine='openpyxl') as writer:
                    df_accordi.to_excel(writer, index=False, sheet_name="Accordi_Nazionali")
                    
                st.download_button(
                    label="Scarica Template Nazionali (Excel)",
                    data=buffer_export.getvalue(),
                    file_name=f"Backup_Contratti_Nazionali_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        with col_b2:
            with st.container(border=True):
                st.markdown("#### Importazione Massiva Contratti Nazionali")
                uploaded_file = st.file_uploader("Trascina il file Excel Contratti (.xlsx)", type=["xlsx"])
                
                if uploaded_file is not None:
                    if st.button("Conferma Scrittura Contratti Nazionali"):
                        try:
                            df_import = pd.read_excel(uploaded_file)
                            colonne_obbligatorie = ["GRUPPO_MACRO", "SOTTOGRUPPO", "LIVELLO", "CHIAVE_LIVELLO"]
                            
                            df_import = DataSanitizer.sanitize_excel_import(df_import, expected_columns=colonne_obbligatorie)
                            
                            cursor = conn.cursor()
                            righe_inserite = 0
                            
                            with conn:
                                for idx, row in df_import.iterrows():
                                    gruppo = str(row["GRUPPO_MACRO"]).upper().strip()
                                    sottogruppo = str(row["SOTTOGRUPPO"]).upper().strip() if (pd.notna(row.get("SOTTOGRUPPO")) and str(row.get("SOTTOGRUPPO")).strip() != "") else ""
                                    livello = str(row["LIVELLO"]).upper().strip()
                                    chiave_livello = str(row["CHIAVE_LIVELLO"]).strip() if pd.notna(row["CHIAVE_LIVELLO"]) else ""
                                    
                                    if livello == "REFERENZA" and chiave_livello:
                                        chiave_livello = str(chiave_livello).split('.')[0].zfill(13)

                                    def to_float_or_none(val):
                                        if pd.isna(val) or str(val).strip() == "": return None
                                        try: return float(val)
                                        except: return None

                                    cursor.execute("""
                                    INSERT OR REPLACE INTO accordi_commerciali (
                                        gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello, listino_r,
                                        sconto_1, sconto_2, sconto_3, sconto_4, sconto_5,
                                        sconto_carico, sconto_pagamento,
                                        voce_contratto_1, voce_contratto_2, voce_contratto_3, voce_contratto_4, voce_contratto_5
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    """, (
                                        gruppo, sottogruppo, "", livello, chiave_livello, to_float_or_none(row.get("LISTINO_BASE_R")),
                                        to_float_or_none(row.get("SCONTO_1")), to_float_or_none(row.get("SCONTO_2")), to_float_or_none(row.get("SCONTO_3")),
                                        to_float_or_none(row.get("SCONTO_4")), to_float_or_none(row.get("SCONTO_5")),
                                        to_float_or_none(row.get("SCONTO_CARICO_LOGISTICA")),
                                        to_float_or_none(row.get("SCONTO_PAGAMENTO_AC")), to_float_or_none(row.get("PFA_VOCE_I")),
                                        to_float_or_none(row.get("PFA_VOCE_II")), to_float_or_none(row.get("PFA_VOCE_III")),
                                        to_float_or_none(row.get("PFA_VOCE_IV")), to_float_or_none(row.get("PFA_VOCE_V"))
                                    ))
                                    righe_inserite += 1
                            st.success(f"Elaborate {righe_inserite} regole commerciali.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Errore durante l'importazione: {e}")

    with tab_guardrail:
        st.markdown("Minimi Net Net €/pz per Referenza")
        with st.container(border=True):
            st.markdown("#### Modifica Diretta Guardrail")
            
            df_guardrail = pd.read_sql_query("""
                SELECT g.ean, a.descrizione_commerciale, g.min_net_net_g
                FROM guardrail_aziendali g
                LEFT JOIN anagrafica_master a ON g.ean = a.ean
            """, conn)
            
            edited_guardrail = st.data_editor(
                df_guardrail, 
                num_rows="dynamic", 
                use_container_width=True,
                hide_index=True,
                disabled=["descrizione_commerciale"],
                column_config={
                    "min_net_net_g": st.column_config.NumberColumn("Min Net Net (€)", format="€ %.2f", step=0.01)
                },
                key="guardrail_editor"
            )
            
            if st.button("SALVA MODIFICHE GUARDRAIL", type="primary"):
                cursor = conn.cursor()
                try:
                    with conn:
                        cursor.execute("DELETE FROM guardrail_aziendali")
                        for _, r in edited_guardrail.iterrows():
                            cursor.execute("INSERT INTO guardrail_aziendali (ean, min_net_net_g) VALUES (?, ?)", (str(r.get("ean")).strip(), float(r.get("min_net_net_g", 0.0))))
                    st.success("Guardrail aggiornati.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Errore salvataggio guardrail: {e}")

        col_g1, col_g2 = st.columns(2)
        
        with col_g1:
            with st.container(border=True):
                st.markdown("#### Esportazione Guardrail")
                
                buffer_guardrail = io.BytesIO()
                with pd.ExcelWriter(buffer_guardrail, engine='openpyxl') as writer:
                    df_guardrail.to_excel(writer, index=False, sheet_name="Guardrail_NetNet")
                    
                st.download_button(
                    label="Scarica Guardrail (Excel)",
                    data=buffer_guardrail.getvalue(),
                    file_name=f"Guardrail_Minimi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        with col_g2:
            with st.container(border=True):
                st.markdown("#### Importazione Massiva Guardrail")
                uploaded_guardrail = st.file_uploader("Trascina il file Excel Guardrail (.xlsx)", type=["xlsx"], key="up_guardrail")
                
                if uploaded_guardrail is not None:
                    if st.button("Conferma Scrittura Guardrail"):
                        try:
                            df_g_import = pd.read_excel(uploaded_guardrail)
                            df_g_import.rename(columns=lambda x: str(x).upper().strip(), inplace=True)
                            df_g_import = DataSanitizer.sanitize_excel_import(df_g_import)
                            df_g_import.columns = [str(c).lower().strip() for c in df_g_import.columns]
                            
                            if "ean" not in df_g_import.columns or "min_net_net_g" not in df_g_import.columns:
                                st.error("Il file Excel deve contenere le colonne 'ean' e 'min_net_net_g'.")
                            else:
                                cursor = conn.cursor()
                                righe_inserite = 0
                                
                                with conn:
                                    for idx, row in df_g_import.iterrows():
                                        ean_val = str(row.get("ean", "")).split('.')[0].zfill(13)
                                        if not ean_val or ean_val == "0000000000000" or ean_val == "nan": continue
                                            
                                        min_g = row.get("min_net_net_g", 0.0)
                                        try: min_g = float(str(min_g).replace(',', '.'))
                                        except: min_g = 0.0
                                            
                                        cursor.execute("INSERT OR REPLACE INTO guardrail_aziendali (ean, min_net_net_g) VALUES (?, ?)", (ean_val, min_g))
                                        righe_inserite += 1
                                        
                                st.success(f"Aggiornati {righe_inserite} limiti minimi.")
                                st.rerun()
                        except Exception as e:
                            st.error(f"Errore durante l'importazione: {e}")

    conn.close()

# ==========================================
# SCHEDA 3.1: ACCORDI LOCALI (PROMO)
# ==========================================
elif menu == "Accordi Locali (Promo)":
    st.title("Accordi Locali (Insegne / Associati)")
    st.markdown("In questa sezione puoi definire gli sconti locali (S6, S7, Y) validi solo per le singole insegne sul territorio. Questi sconti si sommeranno a quelli strutturali del Contratto Nazionale.")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    df_locali = pd.read_sql_query("""
        SELECT a.id, a.associato_insegna, a.livello, 
               COALESCE(m.descrizione_commerciale || ' [' || a.chiave_livello || ']', a.chiave_livello) as prodotto_ean,
               a.sconto_6, a.sconto_7, a.sconto_y, a.note_locali
        FROM accordi_commerciali a
        LEFT JOIN anagrafica_master m ON a.chiave_livello = m.ean
        WHERE a.associato_insegna != '' AND a.associato_insegna IS NOT NULL
    """, conn)
    
    cursor.execute("SELECT DISTINCT associato_insegna FROM struttura_gdo WHERE attivo=1 ORDER BY associato_insegna")
    insegne_attive = [r[0] for r in cursor.fetchall()]
    
    cursor.execute("SELECT ean, descrizione_commerciale FROM anagrafica_master")
    prod_list = [f"{r[1]} [{r[0]}]" for r in cursor.fetchall()]
    
    with st.container(border=True):
        st.markdown("#### Gestione Accordi Locali")
        edited_locali = st.data_editor(
            df_locali,
            num_rows="dynamic",
            hide_index=True,
            use_container_width=True,
            column_config={
                "id": None,
                "associato_insegna": st.column_config.SelectboxColumn("Insegna Locale", options=insegne_attive, required=True),
                "livello": st.column_config.SelectboxColumn("Livello", options=["CATEGORIA", "REFERENZA"], required=True),
                "prodotto_ean": st.column_config.SelectboxColumn("Referenza / Categoria", options=prod_list, required=True),
                "sconto_6": st.column_config.NumberColumn("Sconto 6 (%)", format="%.2f"),
                "sconto_7": st.column_config.NumberColumn("Sconto 7 (%)", format="%.2f"),
                "sconto_y": st.column_config.NumberColumn("Sconto Y (%)", format="%.2f"),
                "note_locali": st.column_config.TextColumn("Note")
            }
        )
        
        if st.button("SALVA ACCORDI LOCALI", type="primary"):
            try:
                with conn:
                    cursor.execute("DELETE FROM accordi_commerciali WHERE associato_insegna != '' AND associato_insegna IS NOT NULL")
                    for _, r in edited_locali.iterrows():
                        def check_nan(val):
                            return float(val) if (pd.notna(val) and str(val).strip() != "") else None
                            
                        raw_chiave = str(r.get("prodotto_ean"))
                        if "[" in raw_chiave and "]" in raw_chiave:
                            chiave_pulita = raw_chiave.split("[")[-1].replace("]", "").strip()
                        else:
                            chiave_pulita = raw_chiave.strip()
                        
                        ins_locale = str(r.get("associato_insegna")).upper().strip() if pd.notna(r.get("associato_insegna")) else ""
                        cursor.execute("SELECT gruppo_macro, sottogruppo FROM struttura_gdo WHERE UPPER(TRIM(associato_insegna)) = ? LIMIT 1", (ins_locale,))
                        res_gdo = cursor.fetchone()
                        
                        g_macro = res_gdo[0] if res_gdo else ""
                        s_grup = res_gdo[1] if (res_gdo and res_gdo[1]) else ""
                            
                        cursor.execute("""
                        INSERT OR REPLACE INTO accordi_commerciali (
                            gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello,
                            sconto_6, sconto_7, sconto_y, note_locali
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            g_macro,
                            s_grup,
                            str(r.get("associato_insegna")).upper().strip() if pd.notna(r.get("associato_insegna")) else "",
                            str(r.get("livello")).upper().strip() if pd.notna(r.get("livello")) else "REFERENZA",
                            chiave_pulita,
                            check_nan(r.get("sconto_6")), check_nan(r.get("sconto_7")), check_nan(r.get("sconto_y")),
                            str(r.get("note_locali")).strip() if pd.notna(r.get("note_locali")) else ""
                        ))
                st.success("Accordi Locali aggiornati correttamente.")
                st.rerun()
            except Exception as e:
                st.error(f"Errore durante l'elaborazione: {e}")
                
    col_l1, col_l2 = st.columns(2)
    with col_l1:
        with st.container(border=True):
            st.markdown("#### Esporta Accordi Locali")
            buffer_locali = io.BytesIO()
            with pd.ExcelWriter(buffer_locali, engine='openpyxl') as writer:
                df_locali.to_excel(writer, index=False, sheet_name="Accordi_Locali")
            st.download_button("Scarica Accordi Locali (Excel)", buffer_locali.getvalue(), f"Accordi_Locali_{datetime.now().strftime('%Y%m%d')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
    with col_l2:
        with st.container(border=True):
            st.markdown("#### Importa Accordi Locali")
            up_locali = st.file_uploader("Carica Excel Accordi Locali", type=['xlsx'], key="up_locali")
            if up_locali:
                if st.button("Conferma Importazione Locali"):
                    try:
                        df_imp_locali = pd.read_excel(up_locali)
                        df_imp_locali = DataSanitizer.sanitize_excel_import(df_imp_locali, expected_columns=["associato_insegna", "livello", "prodotto_ean"])
                        with conn:
                            for _, r in df_imp_locali.iterrows():
                                raw_chiave = str(r.get("prodotto_ean"))
                                if "[" in expandable_check and "]" in raw_chiave:
                                    chiave_pulita = raw_chiave.split("[")[-1].replace("]", "").strip()
                                else:
                                    chiave_pulita = raw_chiave.strip()
                                    
                                def check_nan(val):
                                    return float(val) if (pd.notna(val) and str(val).strip() != "") else None
                                
                                ins_locale = str(r.get("associato_insegna")).upper().strip()
                                cursor.execute("SELECT gruppo_macro, sottogruppo FROM struttura_gdo WHERE UPPER(TRIM(associato_insegna)) = ? LIMIT 1", (ins_locale,))
                                res_gdo = cursor.fetchone()
                                
                                g_macro = res_gdo[0] if res_gdo else ""
                                s_grup = res_gdo[1] if (res_gdo and res_gdo[1]) else ""
                                    
                                cursor.execute("""
                                INSERT OR REPLACE INTO accordi_commerciali (
                                    gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello,
                                    sconto_6, sconto_7, sconto_y, note_locali
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """, (
                                    g_macro,
                                    s_grup,
                                    str(r.get("associato_insegna")).upper().strip(),
                                    str(r.get("livello")).upper().strip(),
                                    chiave_pulita,
                                    check_nan(r.get("sconto_6")), check_nan(r.get("sconto_7")), check_nan(r.get("sconto_y")),
                                    str(r.get("note_locali")).strip() if pd.notna(r.get("note_locali")) else ""
                                ))
                        st.success("Importazione completata.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Errore importazione: {e}")
                        
    conn.close()

# ==========================================
# SCHEDA 4: REPORT SINTETICO E DASHBOARD
# ==========================================
elif menu == "Report Sintetico":
    st.title("Report Sintetico e Analisi Contratti")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # --- 1. POPOLAMENTO DATI ACCOPPIAMENTO COMMERCIALE (df_dash) ---
    cursor.execute("SELECT gruppo_macro, associato_insegna FROM struttura_gdo WHERE attivo=1")
    all_clients = cursor.fetchall()
    
    cursor.execute("SELECT ean, tipo_olio, descrizione_commerciale, COALESCE(min_net_net_g, 0) FROM anagrafica_master LEFT JOIN guardrail_aziendali USING(ean)")
    all_products = cursor.fetchall()
    
    dash_data = []
    for c in all_clients:
        cursor.execute("SELECT sottogruppo FROM accordi_commerciali WHERE gruppo_macro=? AND sottogruppo != '' LIMIT 1", (c[0],))
        res_sub = cursor.fetchone()
        sg = res_sub[0] if res_sub else ""
        
        for p in all_products:
            res = get_merged_contract(conn, c[0], sg, c[1], p[0], p[1])
                
            if res.listino_r is not None:
                p_net = float(res.listino_r)
                for s in [res.sconto_1, res.sconto_2, res.sconto_3, res.sconto_4, res.sconto_5, res.sconto_6, res.sconto_7, res.sconto_y, res.sconto_carico, res.sconto_pagamento]:
                    if s is not None:
                        p_net *= (1 - (float(s)/100))
                pfa = float((res.voce_i or 0) + (res.voce_ii or 0) + (res.voce_iii or 0) + (res.voce_iv or 0) + (res.voce_v or 0))
                p_net *= (1 - (pfa/100))
                
                dash_data.append({
                    'Gruppo': c[0],
                    'Cliente': c[1] if c[1] else c[0],
                    'Categoria': p[1],
                    'Prodotto': p[2],
                    'NetNet': p_net,
                    'Floor': float(p[3]),
                    'Delta_Euro': p_net - float(p[3]),
                    'PFA_Tot': pfa,
                    'Stato': 'Verde (Sopra Soglia)' if p_net >= float(p[3]) else 'Rosso (Sotto Soglia)'
                })
                
    df_dash = pd.DataFrame(dash_data)
    
    # --- 2. KPI CARDS SUPERIORI (Sempre Visibili) ---
    col_k1, col_k2, col_k3, col_k4 = st.columns(4)
    cursor.execute("SELECT COUNT(*) FROM accordi_commerciali")
    col_k1.metric("Totale Regole Attive", f"{cursor.fetchone()[0]}")
    
    cursor.execute("SELECT COUNT(*) FROM struttura_gdo WHERE attivo=1")
    col_k2.metric("Insegne Attive", f"{cursor.fetchone()[0]}")
    
    cursor.execute("SELECT AVG(listino_r) FROM accordi_commerciali WHERE listino_r IS NOT NULL AND listino_r > 0")
    avg_listino = cursor.fetchone()[0] or 0.0
    col_k3.metric("Listino Medio R", fmt_it(avg_listino, is_euro=True))
    
    cursor.execute("""
        SELECT AVG(voce_contratto_1 + COALESCE(voce_contratto_2,0) + COALESCE(voce_contratto_3,0) + 
                   COALESCE(voce_contratto_4,0) + COALESCE(voce_contratto_5,0)) 
        FROM accordi_commerciali
    """)
    avg_pfa = cursor.fetchone()[0] or 0.0
    col_k4.metric("PFA Medio off-invoice", fmt_it(avg_pfa, is_pct=True))
    
    st.divider()

    # ----------------------------------------------------
    # SEZIONE 1: REPORT CONSOLIDATO GRUPPO O INSEGNA
    # ----------------------------------------------------
    st.markdown("### 📋 Report Consolidato Gruppo o Insegna")
    st.markdown("Consolida e verifica l'incolumità finanziaria del portafoglio completo di tutte le referenze e relativi scarti sotto floor.")
    
    with st.container(border=True):
        cursor.execute("SELECT DISTINCT gruppo_macro FROM struttura_gdo WHERE attivo=1 ORDER BY gruppo_macro")
        gruppi_report = [r[0] for r in cursor.fetchall()]
        
        col_rep1, col_rep2, col_rep3 = st.columns(3)
        with col_rep1:
            grp_rep_sel = st.selectbox("1. Gruppo GDO Nazionale", gruppi_report, key="rep_grp")
            
        with col_rep2:
            cursor.execute("""
                SELECT DISTINCT sottogruppo FROM accordi_commerciali WHERE gruppo_macro=? AND sottogruppo != '' AND sottogruppo IS NOT NULL
                UNION
                SELECT DISTINCT sottogruppo FROM struttura_gdo WHERE gruppo_macro=? AND sottogruppo != '' AND sottogruppo IS NOT NULL
                ORDER BY sottogruppo
            """, (grp_rep_sel, grp_rep_sel))
            sottogruppi_rep = [r[0] for r in cursor.fetchall()]
            sottogruppi_rep_options = ["— Nessuno / Solo Gruppo —"] + sottogruppi_rep
            sub_rep_sel = st.selectbox("2. Sottogruppo GDO (Opzionale)", sottogruppi_rep_options, key="rep_sub")
            
        with col_rep3:
            sub_val_query = None if sub_rep_sel == "— Nessuno / Solo Gruppo —" else sub_rep_sel
            if sub_val_query:
                cursor.execute("""
                    SELECT DISTINCT associato_insegna FROM struttura_gdo 
                    WHERE gruppo_macro=? AND sottogruppo=? AND attivo=1 AND associato_insegna != '' AND associato_insegna IS NOT NULL
                    ORDER BY associato_insegna
                """, (grp_rep_sel, sub_val_query))
            else:
                cursor.execute("""
                    SELECT DISTINCT associato_insegna FROM struttura_gdo 
                    WHERE gruppo_macro=? AND attivo=1 AND associato_insegna != '' AND associato_insegna IS NOT NULL
                    ORDER BY associato_insegna
                """, (grp_rep_sel,))
            associati_report = [r[0] for r in cursor.fetchall()]
            associati_report_options = ["— Solo Contratto Strutturale (Nessuna Insegna) —"] + associati_report
            ass_rep_sel = st.selectbox("3. Insegna Locale (Opzionale)", associati_report_options, key="rep_ass")

        st.markdown("<div style='margin-top: 10px;'></div>", unsafe_allow_html=True)
        btn_generate_report = st.button("🔍 GENERA REPORT CONSOLIDATO", type="primary", use_container_width=True)

    # Elaborazione Report su Richiesta
    if btn_generate_report:
            sub_val = None if sub_rep_sel == "— Nessuno / Solo Gruppo —" else sub_rep_sel
            ins_val = None if ass_rep_sel == "— Solo Contratto Strutturale (Nessuna Insegna) —" else ass_rep_sel
            
            cursor.execute("""
                SELECT a.ean, a.descrizione_commerciale, a.tipo_olio, COALESCE(g.min_net_net_g, 0.0), a.codice_sap, a.formato_lt, a.confezione 
                FROM anagrafica_master a
                LEFT JOIN guardrail_aziendali g ON a.ean = g.ean
            """)
            all_prods = cursor.fetchall()
            
            rows_report = []
            for p in all_prods:
                p_ean, p_desc, p_tipo, p_min_g, p_sap, p_form, p_conf = p
                resolved = get_merged_contract(conn, grp_rep_sel, sub_val, ins_val, p_ean, p_tipo)
                
                if resolved.listino_r is None:
                    resolved.listino_r = get_listino_strutturale(conn, grp_rep_sel, sub_val, p_ean)
                
                if resolved.listino_r is not None:
                    input_calc = PricingInput(
                        listino_r=safe_dec(resolved.listino_r), sconto_1=safe_dec(resolved.sconto_1), sconto_2=safe_dec(resolved.sconto_2), sconto_3=safe_dec(resolved.sconto_3),
                        sconto_4=safe_dec(resolved.sconto_4), sconto_5=safe_dec(resolved.sconto_5), sconto_6=safe_dec(resolved.sconto_6), sconto_7=safe_dec(resolved.sconto_7),
                        sconto_y=safe_dec(resolved.sconto_y), sconto_z=Decimal("0.00"), sconto_aa=Decimal("0.00"),
                        sconto_carico=safe_dec(resolved.sconto_carico), sconto_pagamento=safe_dec(resolved.sconto_pagamento),
                        voce_i=safe_dec(resolved.voce_i), voce_ii=safe_dec(resolved.voce_ii), voce_iii=safe_dec(resolved.voce_iii), voce_iv=safe_dec(resolved.voce_iv), voce_v=safe_dec(resolved.voce_v),
                        min_net_net_g=safe_dec(p_min_g)
                    )
                    res_calc = PricingEngine.calculate(input_calc)
                    
                    rows_report.append({
                        "EAN": p_ean,
                        "Codice SAP": p_sap,
                        "Descrizione Prodotto": p_desc,
                        "Listino Base R": float(resolved.listino_r),
                        "Sconti Centrali (S1-S5)": f"{float(resolved.sconto_1 or 0):.1f}% / {float(resolved.sconto_2 or 0):.1f}% / {float(resolved.sconto_3 or 0):.1f}% / {float(resolved.sconto_4 or 0):.1f}% / {float(resolved.sconto_5 or 0):.1f}%",
                        "Sconti Locali (S6/S7/Y)": f"S6: {float(resolved.sconto_6 or 0):.1f}% | S7: {float(resolved.sconto_7 or 0):.1f}% | Y: {float(resolved.sconto_y or 0):.1f}%",
                        "Oneri (AB/AC)": f"AB: {float(resolved.sconto_carico or 0):.1f}% / AC: {float(resolved.sconto_pagamento or 0):.1f}%",
                        "Premi PFA (AL %)": f"{float(res_calc.contratto_tot_pfa):.1f}%",
                        "Prezzo Net Net AM": float(res_calc.net_net_finale),
                        "Soglia Floor G": float(p_min_g),
                        "Delta Margine": float(res_calc.delta_vs_min),
                        "Stato Approvazione": "🟢 Approvato" if res_calc.guardrail_ok else "🔴 Sotto Floor"
                    })
            
            if not rows_report:
                st.warning("Nessun prodotto o accordo commerciale attivo per i criteri selezionati.")
            else:
                df_rep_out = pd.DataFrame(rows_report)
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Report_Consolidato"
                
                font_header = Font(name="Space Grotesk", size=11, bold=True, color="FFFFFF")
                fill_header = PatternFill(start_color="5A6340", end_color="5A6340", fill_type="solid")
                fill_red = PatternFill(start_color="FAF2F0", end_color="FAF2F0", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin', color='E2E2D8'), right=Side(style='thin', color='E2E2D8'),
                    top=Side(style='thin', color='E2E2D8'), bottom=Side(style='thin', color='E2E2D8')
                )
                
                for col_num, h_text in enumerate(df_rep_out.columns, 1):
                    cell = ws.cell(row=1, column=col_num, value=h_text)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                for row_num, row_data in enumerate(df_rep_out.values, 2):
                    is_red = "🔴" in str(row_data[-1])
                    for col_num, val in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=val)
                        cell.border = thin_border
                        cell.font = Font(name="Inter", size=10)
                        if is_red:
                            cell.fill = fill_red
                        
                        # Riallineamento indici valute dopo inserimento nuova colonna (4, 9, 10, 11)
                        if col_num in [4, 9, 10, 11]:
                            cell.number_format = '#,##0.000 €' if col_num in [9, 11] else '#,##0.00 €'
                
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                    
                buffer_rep = io.BytesIO()
                wb.save(buffer_rep)
                
                st.session_state.compiled_rep_df = df_rep_out
                st.session_state.compiled_rep_bytes = buffer_rep.getvalue()
                nome_file_exp = f"Report_Consolidato_{grp_rep_sel.replace(' ', '_')}"
                if sub_val:
                    nome_file_exp += f"_{sub_val.replace(' ', '_')}"
                if ins_val:
                    nome_file_exp += f"_{ins_val.replace(' ', '_')}"
                st.session_state.compiled_rep_filename = f"{nome_file_exp}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    if "compiled_rep_df" in st.session_state:
        st.markdown("#### 📄 Risultati Report Consolidato Elaborato")
        
        st.download_button(
            label="📥 SCARICA REPORT EXCEL (.xlsx)",
            data=st.session_state.compiled_rep_bytes,
            file_name=st.session_state.compiled_rep_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        st.dataframe(
            st.session_state.compiled_rep_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Listino Base R": st.column_config.NumberColumn("Listino Base R", format="€ %.2f"),
                "Prezzo Net Net AM": st.column_config.NumberColumn("Prezzo Net Net AM", format="€ %.3f"),
                "Soglia Floor G": st.column_config.NumberColumn("Soglia Floor G", format="€ %.2f"),
                "Delta Margine": st.column_config.NumberColumn("Delta Margine", format="€ %+.3f"),
                "Stato Approvazione": st.column_config.TextColumn("Stato Approvazione"),
                "EAN": st.column_config.TextColumn("EAN"),
                "Codice SAP": st.column_config.TextColumn("Codice SAP")
            }
        )
    
    st.divider()

    # --- 3. SEZIONE 2: BENCHMARK COMPARATIVO DI CANALE ---
    st.markdown("### 🔍 Benchmark Comparativo di Canale (Livello Sottogruppo)")
    st.markdown("Analisi strutturale delle asimmetrie commerciali. Sconti e oneri collassati per destinazione logica.")
    
    with st.container(border=True):
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            cursor.execute("SELECT DISTINCT tipo_olio FROM anagrafica_master ORDER BY tipo_olio")
            categorie_disponibili = [r[0] for r in cursor.fetchall()]
            cat_scelta = st.selectbox("Filtra per Categoria Merceologica", categorie_disponibili, key="bench_cat_old")
            
        with col_f2:
            cursor.execute("SELECT ean, descrizione_commerciale FROM anagrafica_master WHERE tipo_olio=? ORDER BY descrizione_commerciale", (cat_scelta,))
            prod_dict = {f"{p[1]} [{p[0]}]": (p[0], cat_scelta) for p in cursor.fetchall()}
            if prod_dict:
                prod_scelto_bench = st.selectbox("Seleziona Referenza da Analizzare", list(prod_dict.keys()), key="bench_prod_old")
                ean_bench, tipo_olio_bench = prod_dict[prod_scelto_bench]
            else:
                st.warning("Nessun prodotto trovato.")
                prod_scelto_bench = None

        if prod_scelto_bench:
            cursor.execute("""
                SELECT DISTINCT gruppo_macro, sottogruppo 
                FROM accordi_commerciali 
                WHERE sottogruppo != '' AND sottogruppo IS NOT NULL
                ORDER BY gruppo_macro, sottogruppo
            """)
            sottogruppi_unici = cursor.fetchall()
            benchmark_data = []
            
            for g_macro, s_gruppo in sottogruppi_unici:
                cursor.execute("""
                    SELECT associato_insegna FROM accordi_commerciali
                    WHERE gruppo_macro=? AND sottogruppo=? AND livello='REFERENZA' AND chiave_livello=? AND associato_insegna != ''
                    LIMIT 1
                """, (g_macro, s_gruppo, ean_bench))
                res_ins = cursor.fetchone()
                
                if not res_ins:
                    cursor.execute("""
                        SELECT associato_insegna FROM accordi_commerciali
                        WHERE gruppo_macro=? AND sottogruppo=? AND associato_insegna != ''
                        LIMIT 1
                    """, (g_macro, s_gruppo))
                    res_ins = cursor.fetchone()
                
                insegna_campione = res_ins[0] if res_ins else ""
                
                contratto_risolto = get_merged_contract(conn, g_macro, s_gruppo, insegna_campione, ean_bench, tipo_olio_bench)
                
                if contratto_risolto.listino_r is not None:
                    cursor.execute("SELECT min_net_net_g FROM guardrail_aziendali WHERE ean=?", (ean_bench,))
                    res_g = cursor.fetchone()
                    soglia_g = res_g[0] if res_g else 0.0
                    
                    input_strutturale = PricingInput(
                        listino_r=safe_dec(contratto_risolto.listino_r),
                        sconto_1=safe_dec(contratto_risolto.sconto_1), sconto_2=safe_dec(contratto_risolto.sconto_2), sconto_3=safe_dec(contratto_risolto.sconto_3),
                        sconto_4=safe_dec(contratto_risolto.sconto_4), sconto_5=safe_dec(contratto_risolto.sconto_5), sconto_6=safe_dec(contratto_risolto.sconto_6), sconto_7=safe_dec(contratto_risolto.sconto_7),
                        sconto_y=safe_dec(contratto_risolto.sconto_y), sconto_z=Decimal("0.00"), sconto_aa=Decimal("0.00"),
                        sconto_carico=safe_dec(contratto_risolto.sconto_carico), sconto_pagamento=safe_dec(contratto_risolto.sconto_pagamento),
                        voce_i=safe_dec(contratto_risolto.voce_i), voce_ii=safe_dec(contratto_risolto.voce_ii), voce_iii=safe_dec(contratto_risolto.voce_iii), voce_iv=safe_dec(contratto_risolto.voce_iv), voce_v=safe_dec(contratto_risolto.voce_v),
                        min_net_net_g=safe_dec(soglia_g)
                    )
                    calcolo_strutturale = PricingEngine.calculate(input_strutturale)
                    
                    stringa_s1_s3 = f"{float(contratto_risolto.sconto_1 or 0):.1f}% / {float(contratto_risolto.sconto_2 or 0):.1f}% / {float(contratto_risolto.sconto_3 or 0):.1f}%"
                    stringa_s4_s5 = f"{float(contratto_risolto.sconto_4 or 0):.1f}% / {float(contratto_risolto.sconto_5 or 0):.1f}%"
                    stringa_s6 = f"{float(contratto_risolto.sconto_6 or 0):.1f}%"
                    stringa_s7_y = f"S7:{float(contratto_risolto.sconto_7 or 0):.1f}% + Y:{float(contratto_risolto.sconto_y or 0):.1f}%"
                    stringa_oneri = f"Log:{float(contratto_risolto.sconto_carico or 0):.1f}% / Pag:{float(contratto_risolto.sconto_pagamento or 0):.1f}%"
                    
                    benchmark_data.append({
                        "Gruppo GDO": g_macro,
                        "Sottogruppo": s_gruppo,
                        "Origine Accordo": contratto_risolto.livello_risolto,
                        "Listino R (€)": float(contratto_risolto.listino_r),
                        "Gruppo (S1-S3)": stringa_s1_s3,
                        "Sottogruppo (S4-S5)": stringa_s4_s5,
                        "Categoria (S6)": stringa_s6,
                        "Referenza (S7+Y)": stringa_s7_y,
                        "Oneri (AB/AC)": stringa_oneri,
                        "Contratto Unificato (%)": float(calcolo_strutturale.contratto_tot_pfa),
                        "Net Net Base AM (€)": float(calcolo_strutturale.net_net_finale)
                    })
            
            if benchmark_data:
                df_out = pd.DataFrame(benchmark_data).sort_values(by="Net Net Base AM (€)")
                st.dataframe(df_out, use_container_width=True, hide_index=True)
            else:
                st.info("Nessun accordo strutturato trovato per i filtri selezionati.")

        st.divider()

        # --- 4. SEZIONE 3: SINTESI CANALE E DASHBOARD DIREZIONALE ---
        st.markdown("### 📊 Salute Contratti, Profondità Margine & Sintesi Canale GDO")
        
        if not df_dash.empty:
            col_sin_left, col_sin_right = st.columns([1, 1])
            
            with col_sin_left:
                with st.container(border=True):
                    st.markdown("#### Sintesi Dinamica per Canale GDO")
                    query_sintesi = """
                        SELECT gruppo_macro as [Gruppo Macro],
                               COUNT(*) as [Totale Righe],
                               ROUND(AVG(listino_r), 2) as [Listino Medio (€)],
                               ROUND(AVG(sconto_1), 2) as [Sconto 1 Medio (%)],
                               ROUND(AVG(voce_contratto_1 + COALESCE(voce_contratto_2,0) + COALESCE(voce_contratto_3,0) + 
                                         COALESCE(voce_contratto_4,0) + COALESCE(voce_contratto_5,0)), 2) as [PFA Totale (%)]
                        FROM accordi_commerciali
                        GROUP BY gruppo_macro
                        ORDER BY [Totale Righe] DESC
                    """
                    df_sintesi = pd.read_sql_query(query_sintesi, conn)
                    st.dataframe(df_sintesi, use_container_width=True, hide_index=True)
                    
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown("##### Pressione Promozionale (PFA) per Categoria")
                    df_cat_health = df_dash.groupby('Categoria')['PFA_Tot'].mean().reset_index().sort_values('PFA_Tot', ascending=False)
                    fig_cat = px.bar(df_cat_health, x='Categoria', y='PFA_Tot', labels={'Categoria': 'Famiglia Prodotto', 'PFA_Tot': 'PFA Medio (%)'}, color='PFA_Tot', color_continuous_scale='Blues')
                    st.plotly_chart(fig_cat, use_container_width=True)

            with col_sin_right:
                with st.container(border=True):
                    st.markdown("#### Salute dei Contratti")
                    
                    col_chart_f1, col_chart_f2 = st.columns(2)
                    with col_chart_f1:
                        dash_cat = st.selectbox("1. Filtra Grafici per Categoria", ["Tutte le Categorie"] + sorted(df_dash['Categoria'].unique().tolist()), key="dash_cat")
                    with col_chart_f2:
                        prods_available = df_dash[df_dash['Categoria'] == dash_cat]['Prodotto'].unique().tolist() if dash_cat != "Tutte le Categorie" else df_dash['Prodotto'].unique().tolist()
                        dash_prod = st.selectbox("2. Filtra Grafici per Referenza", ["Tutte le Referenze"] + sorted(prods_available), key="dash_prod")
                    
                    df_pie = df_dash.copy()
                    if dash_cat != "Tutte le Categorie": 
                        df_pie = df_pie[df_pie['Categoria'] == dash_cat]
                    if dash_prod != "Tutte le Referenze": 
                        df_pie = df_pie[df_pie['Prodotto'] == dash_prod]
                    
                    if not df_pie.empty:
                        def format_clients(clients):
                            unique_clients = sorted(list(set(clients)))
                            if len(unique_clients) > 8:
                                return "<br>".join(unique_clients[:8]) + "<br><i>...e altri</i>"
                            return "<br>".join(unique_clients)

                        df_pie_agg = df_pie.groupby('Stato').agg(
                            Conteggio=('Prodotto', 'count'),
                            Clienti_Lista=('Cliente', format_clients)
                        ).reset_index()

                        fig_pie = px.pie(df_pie_agg, values='Conteggio', names='Stato',
                                         custom_data=['Clienti_Lista'],
                                         title="Distribuzione Referenze (Sopra/Sotto Soglia Vs net net contrattuale)",
                                         color='Stato', color_discrete_map={'Verde (Sopra Soglia)':'#5A6340', 'Rosso (Sotto Soglia)':'#A34A3F'})
                        
                        fig_pie.update_traces(hovertemplate="<b>%{label}</b><br>Num. Accordi: %{value}<br><br><b>Clienti coinvolti:</b><br>%{customdata[0]}<extra></extra>")
                        st.plotly_chart(fig_pie, use_container_width=True)
                        
                        df_delta = df_pie.groupby('Categoria')['Delta_Euro'].mean().reset_index()
                        df_delta['Colore'] = df_delta['Delta_Euro'].apply(lambda x: 'Positivo' if x >= 0 else 'Negativo')
                        
                        fig_delta = px.bar(df_delta, x='Categoria', y='Delta_Euro', 
                                           title="Distanza Media dal Floor (€)",
                                           color='Colore', color_discrete_map={'Positivo':'#5A6340', 'Negativo':'#A34A3F'},
                                           labels={'Delta_Euro': 'Delta Medio (€)', 'Categoria': ''})
                        
                        fig_delta.update_layout(showlegend=False)
                        fig_delta.update_traces(hovertemplate="<b>%{x}</b><br>Delta Medio: %{y:.3f} €<extra></extra>")
                        st.plotly_chart(fig_delta, use_container_width=True)
                        
                    else:
                        st.info("Nessun dato disponibile per i filtri selezionati.")
        else:
            st.warning("Nessun contratto attivo trovato nel database per generare la Dashboard.")

    conn.close()
# ==========================================
# SCHEDA 5: GUIDA OPERATIVA (MANUALE UTENTE COMPLETO)
# ==========================================
else:
    st.title("📚 Manuale Operativo e Linee Guida")
    st.markdown("### Guida all'uso di 'Bunker Commerciale - Salov'")
    st.markdown("Benvenuto nel sistema aziendale per il governo della marginalità. Questa guida spiega le logiche di calcolo del software e fornisce le istruzioni passo-passo per condurre le simulazioni commerciali in totale sicurezza.")
    
    st.divider()

    with st.expander("📖 1. GLOSSARIO: I termini tecnici da conoscere", expanded=True):
        st.markdown("""
        Per utilizzare correttamente il simulatore, è fondamentale condividere lo stesso vocabolario tecnico:

        *   **Listino Base (R):** È il prezzo di listino ufficiale Salov, al lordo di qualsiasi sconto.
        *   **Sconti in Fattura (S1... S7, Y, Z):** Sconti percentuali applicati direttamente in fattura. Riducono l'imponibile. Si dividono in *Centrali* (definiti dall'Accordo Quadro) e *Locali/Promo* (definiti per singole attività).
        *   **Sconto Diretto (AA):** È un "Taglio Prezzo" espresso in Valore Absoluto (Euro), non in percentuale (es. -0,50 € a bottiglia).
        *   **Oneri Logistici (AB) e di Pagamento (AC):** Trattenute percentuali applicate dal cliente per la gestione centralizzata del magazzino o per i flussi finanziari.
        *   **Netto Fattura 2 (AF):** Il prezzo reale a cui il prodotto viene fatturato, calcolato dopo aver applicato tutti gli sconti e gli oneri.
        *   **Premi Fuori Fattura / Off-Invoice (PFA):** (Voci I, II, III, ecc.). Sono i contributi di fine anno o fine periodo richiesti dalla GDO (es. premi di fine anno, contributi assortimento). *Riducono il nostro margine, ma non abbassano il prezzo a scaffale del cliente*.
        *   **NET-NET FINALE (AM):** È il ricavo reale e pulito per l'azienda. Si ottiene sottraendo i PFA dal Netto Fattura.
        *   **Minimo Net-Net:** È la soglia minima di redditività fissata dall'azienda per una specifica referenza. Scendere sotto questo valore significa vendere in perdita.
        *   **Floor (G) / Guardrail Aziendale:** Rappresenta il limite fisico e invalicabile impostato dalla Direzione nel database, il "pavimento" sotto il quale il sistema blocca l'operazione. Agisce come un vero e proprio salvavita.
            * *Esempio pratico:* Immagina che produrre, imbottigliare e consegnare una bottiglia di Extravergine costi 3,50 €. L'azienda fissa il Floor (G) a 3,80 € per garantirsi la sopravvivenza e coprire i costi fissi. Se durante una negoziazione concedi un mix di sconti, premi di fine anno e contributi volantino che fa crollare il tuo ricavo reale a 3,75 €, il sistema farà scattare il semaforo **ROSSO**. Ti avviserà immediatamente che stai "bucando il Floor", distruggendo valore per 0,05 € su ogni singola bottiglia venduta.
        *   **Spazio Promo MAX (% o €):** Indica il margine di manovra residuo. È la differenza tra il prezzo simulato e il Floor minimo aziendale.
        """)
        
    with st.expander("🧮 2. LA MECCANICA DEGLI SCONTI: La 'Cascata' e il Sell-Out", expanded=False):
        st.markdown("""
        **La Cascata Geometrica (Sconti in Fattura)**
        Il software non fa mai la somma algebrica degli sconti in fattura (es. 10% + 5% NON fa 15%). Il calcolo è **sequenziale**: ogni sconto si calcola sul valore residuo lasciato dallo sconto precedente.

        *Esempio di Calcolo in Fattura:*
        *   **Listino Base:** 10,00 €
        *   **Sconto 1 (10%):** 10,00 * (1 - 0.10) = **9,00 €**
        *   **Sconto 2 (5%):** 9,00 * (1 - 0.05) = **8,55 €**
        *   **Sconto Promo Z (10%):** 8,55 * (1 - 0.10) = **7,695 €**
        *   **Sconto Diretto AA (-0,20 €):** 7,695 - 0,20 = **7,495 €**
        *   **Oneri Logistica (2%):** 7,495 * (1 - 0.02) = **7,345 € (Netto Fattura)**

        **La Somma Algebrica (Premi Fuori Fattura - PFA)**
        A differenza degli sconti in fattura, i PFA si **sommano algebricamente** tra loro prima di essere applicati. Se hai un Premio Base del 2% e un Contributo Volantino dell'1%, il sistema calcolerà un 3% totale sul Netto Fattura.
        *   *Calcolo Net-Net:* 7,345 € * (1 - 0.03) = **7,124 € (Net-Net Finale)**.

        **I Contributi Extra (Sell-Out / Volantino)**
        Se durante una promo concedi un contributo fisso (es. 500 € per una testata gondola) a fronte di 10.000 bottiglie stimate:
        *   Impatto unitario: 500 € / 10.000 pz = **0,05 € a bottiglia**.
        *   Il tuo Net-Net reale scenderà da 7,124 € a **7,074 €**. Il sistema calcola questo impatto per verificare che tu non scenda sotto il Floor minimo.
        """)

    with st.expander("🧬 3. LA GERARCHIA DEI CONTRATTI: La regola del 'Top-Down Lock'", expanded=False):
        st.markdown("""
        Il database contratti funziona con una logica a 5 livelli. Vige la regola del **Blocco dall'Alto (Top-Down Lock)**: se un livello superiore fissa una condizione, i livelli inferiori non possono modificarla o cancellarla.
        
        **I 5 Livelli (dal più forte al più debole):**
        1. **GRUPPO MACRO** (es. *COOP ITALIA*) ➔ Accordo Quadro Nazionale.
        2. **SOTTOGRUPPO** (es. *COOP NORD OVEST*) ➔ Accordi interregionali.
        3. **CATEGORIA** (es. *EXTRAVERGINE*) ➔ Regole valide solo per una specifica famiglia di prodotti.
        4. **ASSOCIATO / INSEGNA** (es. *IPERCOOP LOCALE*) ➔ Accordi del singolo punto vendita o associato.
        5. **REFERENZA (EAN)** ➔ La singola bottiglia. Qui risiede il Listino Base (R).

        **Casi Pratici di Inserimento Dati:**
        *   **L'Ereditarietà (Cella Vuota):** Se il Gruppo Macro ha uno Sconto 1 del 10%, non serve riscriverlo sulle singole referenze. Lasciando la cella vuota, il sistema erediterà automaticamente il 10%.
        *   **L'Override (Forzare l'esclusione):** Se una specifica referenza Premium NON deve ricevere lo Sconto 1 del 10% stabilito dal Gruppo, devi inserire esplicitamente il valore **`0.0`** nella riga di quella referenza. Questo `0.0` funge da scudo e blocca l'ereditarietà.
        *   **Fuori Assortimento:** Se una referenza non ha un Listino Base (R) associato, il sistema la bloccherà indicando "Prodotto Fuori Assortimento".
        """)
        
    with st.expander("📞 4. COME USARE IL 'SIMULATORE OFFERTE' (Singola Referenza)", expanded=False):
        st.markdown("""
        Questa scheda è lo strumento tattico da usare durante una trattativa rapida su un singolo prodotto.
        
        **Modalità A: Partenza da Prezzo Target (Consigliata)**
        Da usare quando il Buyer fissa un obiettivo di prezzo. Esempio: *"Voglio pagare questa bottiglia 3,80 € netti"*.
        1. Seleziona la Modalità A.
        2. Inserisci `3.80` nel campo "Prezzo Target Net Net".
        3. Il sistema calcolerà automaticamente la percentuale esatta di **Sconto Promozionale [Z]** necessaria per arrivare a quel risultato; questo valore varia dinamicamente ma non è modificabile direttamante in quanto va a garantire il rispetto del limite net net pre impostato
        4. Se 3,80 € è inferiore al minimo aziendale (Floor), il sistema mostrerà l'avviso in **ROSSO**, indicando la perdita esatta.
        5. Le uniche leve utilizzabili sono lo Sconto continuativo % (da utilizzare previa verifica di accordo locale) e lo sconto unitario in fattura. Al variare di questi lo Sconto Promozionale si adatterà di conseguenza.

        **Modalità B: Tentativi Spot Manuali**
        Da usare per simulazioni libere.
        1. Inserisci manualmente le percentuali di Sconto Promo [Z] o lo Sconto in Euro [AA].
        2. Controlla i campi **Sconto Promo MAX** e **Sconto Unitario MAX**: ti indicano il limite massimo che puoi concedere prima che il semaforo diventi rosso.
        """)
        
    with st.expander("📊 5. COME USARE LA 'MASTER GRID RINNOVI' (Simulazione N vs N+1)", expanded=False):
        st.markdown("""
        Questa scheda è lo strumento strategico per i rinnovi annuali. Permette di simulare l'intero portafoglio prodotti di un cliente in un'unica schermata, partendo dagli aggregati fino ad arrivare al dettaglio delle singole voci contrattuali.
        
        **Flusso di Lavoro Passo-Passo:**
        
        **FASE 1: Setup e Condizioni Globali**
        1. **Identifica il Cliente:** Seleziona il *Gruppo GDO* e il *Sottogruppo* dai menu a tendina.
        2. **Importa lo Storico:** Clicca su **"Carica Condizioni Attuali da DB"**. Il sistema popolerà le colonne dell'Anno [N] con i listini e gli sconti attualmente in vigore.
        3. **Imposta gli Oneri Globali:** In cima al Tab 1, verifica e imposta lo **Sconto Carico Logistica (%)** e lo **Sconto Pagamento (%)**. *Attenzione: questi due valori vengono applicati a cascata su TUTTE le referenze attive e abbattono direttamente il Net Net finale.*

        **FASE 2: Tab 1 - Griglia di Simulazione (Gli Aggregati)**
        1. **Attiva le Referenze:** Inserisci i volumi previsti nella colonna `[N+1] Volumi`. *Il sistema calcolerà e mostrerà nei risultati SOLO le righe con volumi maggiori di zero.*
        2. **Imposta i Target:** Agisci sulle colonne `[N+1] Listino €`, `[N+1] Sc. Fattura %` (il totale degli sconti in fattura) e `[N+1] Contratto %` (il totale dei premi fuori fattura).
        3. **Calcola:** Clicca su **"🔄 Calcola Simulazione"**.
        4. **Verifica lo Spazio Promo:** Controlla la colonna `Sc. Promo MAX [N+1] %`. Questo valore ti dice quanto sconto volantino potrai fare durante l'anno prima di bucare il limite minimo aziendale (Floor). Modifica i target finché non ottieni lo Spazio Promo desiderato.

        **FASE 3: Tab 2 - Analisi Ponderata (L'Effetto "Pollo di Trilussa")**
        In questa scheda vedi la sintesi economica. I risultati sono raggruppati per Categoria e Sub-Categoria. 
        *Perché è fondamentale?* Un cliente potrebbe avere un margine totale positivo (Verde) perché muove enormi volumi di Olio di Semi, ma nascondere una forte perdita (Rosso) sugli Extravergini. Il sistema calcola la media ponderata sui volumi per evidenziare se un cluster specifico sta distruggendo valore.

        **FASE 4: Tab 3 - Esplosione Sconti (Il Dettaglio Contrattuale)**
        Una volta che i target aggregati nel Tab 1 ti soddisfano, devi "spacchettarli" nelle reali voci del contratto nazionale (S1, S2, PFA I, ecc.).
        1. **Modifica Manuale:** Inserisci i valori noti (es. S1 al 10%, S2 al 2%).
        2. **Verifica:** Clicca su **"🔄 Calcola e Verifica Sconti (Manuale)"**. Il sistema calcolerà la cascata geometrica e ti mostrerà sotto "Diff. Fattura" quanto manca per raggiungere il target che avevi fissato nel Tab 1.
        3. **La Magia (Auto-Allineamento):** Non impazzire con la calcolatrice! Clicca su **"🪄 Allinea Sconti Automaticamente"**. Il sistema calcolerà l'esatta percentuale geometrica mancante e la inserirà in **S5 %**, portando la differenza a zero. Farà la stessa cosa per i premi fuori fattura, inserendo la differenza algebrica in **PFA V %**.
        4. **Trucco:** se vuoi agire su uno dei due sconti in fattura presenti senza doverne creare un terzo (l'S5 della simulazione) azzera uno degli sconti contrattuali, avvia la simulazione e poi sostituisci la colonna dello sconto contrattuale che vuoi modificare con i valori in S5. Ripeti l'allineamento automatico per evidenziare eventuali errori. 
        
        *Nota: In ogni Tab è presente un pulsante per scaricare la tabella corrente in Excel.*
        """)

    with st.expander("🗄️ 6. STORICO (CRM), REPORTISTICA E BACK-OFFICE", expanded=False):
        st.markdown("""
        *   **Storico Promozioni (CRM):** Ogni simulazione può essere salvata nel database (come "Proposta" o "Confermata"). In questa scheda puoi filtrare, consultare ed esportare in Excel tutte le trattative passate. Se hai commesso un errore, puoi eliminare il singolo record tramite il suo ID.
        *   **Clona Promozione:** Nello Storico puoi selezionare una vecchia promo e cliccare "Clona". Il sistema ti riporterà al Simulatore precompilando tutti i campi, permettendoti di creare una nuova trattativa in secondi.
        *   **Report Sintetico:** Genera un file Excel consolidato per un intero cliente. Mostra l'allineamento di tutti i prezzi e sconti, evidenziando immediatamente le referenze approvate (Verdi) e quelle sotto soglia (Rosse). È il documento ideale da condividere con la Direzione Commerciale.
        *   **Back-Office (Import/Export Excel):** Per aggiornare massivamente le anagrafiche, i guardrail (Floor) o i contratti quadro, non occorre farlo riga per riga a schermo.
            1. Scarica il Template Excel.
            2. Modifica i dati sul tuo computer. **ATTENZIONE:** Assicurati che in Excel la colonna degli EAN sia formattata come "Testo", altrimenti Excel trasformerà i codici a barre in numeri scientifici (es. 8,0022E+12) corrompendo il database.
            3. Ricarica il file tramite l'apposito pulsante.
        *   **Danger Zone (Reset):** Il pulsante di Reset nella barra laterale (protetto da password) cancella tutto il database e ricarica i dati finti di test. Da usare solo in fase di training o manutenzione.
        """)
